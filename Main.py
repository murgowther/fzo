from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import *
import psycopg2
from docxtpl import DocxTemplate, R
import pymorphy2
import sys
import pandas as pd
import sqlite3
import asyncio
import time
import openpyxl
from openpyxl import load_workbook

class PageWindow(QtWidgets.QMainWindow):

    gotoSignal = QtCore.pyqtSignal(str)
    def goto(self, name):
        self.gotoSignal.emit(name)
    def con(self):
        self.conn = psycopg2.connect(user="postgres",
                                     password="NlOIBpyp",
                                     host="127.0.0.1",
                                     port="5432",
                                     database="fzo")
        self.cur = self.conn.cursor()
class MainWindow(PageWindow):
    def __init__(self):
        super().__init__()
        self.app = QtWidgets.QApplication(sys.argv)
        self.app.setWindowIcon(QtGui.QIcon('source/герб.ico'))
        self.initUI()
        self.setWindowTitle("Факультет заочного обучения")
        self.setObjectName("MainWindow")
    def initUI(self):
        self.UiComponents()

    def UiComponents(self):
        self.centralwidget = QtWidgets.QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        self.centralwidget.setStyleSheet("background-image:url(source/background.png)")
        self.centralwidget.setFixedSize(1126, 826)
        self.groupBox = QtWidgets.QGroupBox(self)
        self.groupBox.setGeometry(QtCore.QRect(-1, 0, 1141, 112))
        self.groupBox.setStyleSheet("background-color: #343a40;")
        self.groupBox.setTitle("")
        self.groupBox.setAlignment(QtCore.Qt.AlignCenter)
        self.groupBox.setObjectName("groupBox")
        self.groupBox_2 = QtWidgets.QGroupBox(self)
        self.groupBox_2.setEnabled(True)
        self.groupBox_2.setGeometry(QtCore.QRect(390, 320, 361, 251))
        self.groupBox_2.setStyleSheet("background-color: white;\n"
                                      "border-radius: 5;\n"
                                      "border: 2px solid rgb(69, 90, 100);\n"
                                      "border-radius: 10;\n"
                                      "font: 14pt \"Arial\";\n"
                                      "color: white;"
                                      )
        self.groupBox_2.setTitle("")
        self.groupBox_2.setObjectName("groupBox_2")
        self.label_2 = QtWidgets.QLabel("Факультет заочного обучения",self.groupBox_2)
        self.label_2.setGeometry(QtCore.QRect(20, 30, 331, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        self.label_2.setFont(font)
        self.label_2.setStyleSheet("color: rgb(69, 90, 100);\n"
                                   "font: 14pt \"Arial\";\n"
                                   "border: 2px solid white;")
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName("label_2")
        self.searchButton = QtWidgets.QPushButton("Абитуриент",self.groupBox_2)
        self.searchButton.setGeometry(QtCore.QRect(22, 103, 321, 51))
        self.searchButton.setStyleSheet("font: 15pt \"Arial\";\n"
                                      )
        self.searchButton.setStyleSheet("""
            QPushButton:hover { background-color: #bbbbbb }
            QPushButton:!hover { background-color: rgb(69, 90, 100) }
            QPushButton:pressed { background-color: black; }
        """)
        self.searchButton.clicked.connect(
            self.make_handleButton("abiturient")
        )
        self.pushButton = QtWidgets.QPushButton("Слушатель",self.groupBox_2)
        self.pushButton.setGeometry(QtCore.QRect(22, 160, 321, 51))
        self.pushButton.setStyleSheet("""
                    QPushButton:hover { background-color: #bbbbbb }
                    QPushButton:!hover { background-color: rgb(69, 90, 100) }
                    QPushButton:pressed { background-color: black; }
                """)
        self.pushButton.clicked.connect(
            self.make_handleButton("Slushatel")
        )
        self.label_3 = QtWidgets.QLabel(self.groupBox)
        self.label_3.setGeometry(QtCore.QRect(60, 10, 75, 90))
        self.label_3.setMaximumSize(QtCore.QSize(75, 90))
        self.label_3.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label_3.setPixmap(QtGui.QPixmap("source/герб.png"))
        self.label_3.setScaledContents(True)
        self.label_3.setAlignment(QtCore.Qt.AlignCenter)
        self.label_3.setObjectName("label_3")


        self.label = QtWidgets.QLabel("Московский университет МВД России имени В.Я. Кикотя", self.groupBox)
        self.label.setGeometry(QtCore.QRect(200, 40, 850, 38))
        self.label.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label.setStyleSheet("border-color: rgb(255, 255, 255);\n"
                                 "font: 20pt \"Arial\";\n"
                                 "color: rgb(255, 255, 255);")
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")


    def make_handleButton(self, button):
        def handleButton():
            if button == "abiturient":
                self.goto("Abiturient")
            elif button == "Slushatel":
                self.goto("Slushatel")
        return handleButton



class Abiturient(PageWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Абитуриент")
        self.UiComponents()

    def goToMain(self):
        self.goto("main")
    def goToHome(self):
        self.goto("main")
    def UiComponents(self):
        self.setStyleSheet("background-color: white;")
        self.groupBox = QtWidgets.QGroupBox(self)
        self.groupBox.setEnabled(True)
        self.groupBox.setGeometry(QtCore.QRect(-1, 0, 71, 831))
        self.groupBox.setStyleSheet("background-color: rgb(69, 90, 100);")
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.groupBox_2 = QtWidgets.QGroupBox(self)
        self.groupBox_2.setGeometry(QtCore.QRect(100, 20, 1010, 780))
        self.groupBox_2.setStyleSheet("font: 14pt \"Arial\";\n"
                                      "color: white;\n"
                                      "border: 2px solid rgb(69, 90, 100);\n"
                                      "border-radius: 10;\n"
                                      )
        self.backButton = QtWidgets.QPushButton(self.groupBox)
        self.backButton.setGeometry(QtCore.QRect(10, 10, 50, 50))
        self.backButton.setStyleSheet("border-image: url(source/back.png);")
        self.backButton.clicked.connect(self.goToMain)
        self.homeButton = QtWidgets.QPushButton(self.groupBox)
        self.homeButton.setGeometry(QtCore.QRect(10, 70, 50, 50))
        self.homeButton.setStyleSheet("border-image: url(source/home.png);")
        self.homeButton.clicked.connect(self.goToHome)

        self.createButton = QtWidgets.QPushButton("Новый год набора", self.groupBox_2)
        self.createButton.setGeometry(QtCore.QRect(320, 170, 340, 51))
        self.createButton.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb }
                                    QPushButton:!hover { background-color: rgb(69, 90, 100) }
                                    QPushButton:pressed { background-color: black; }
                                """)
        self.createButton.clicked.connect(
            self.make_handleButton("New_Year")
        )
        self.komplektButton = QtWidgets.QPushButton("План комплектования", self.groupBox_2)
        self.komplektButton.setGeometry(QtCore.QRect(320, 250, 340, 51))
        self.komplektButton.setStyleSheet("""
                                            QPushButton:hover { background-color: #bbbbbb }
                                            QPushButton:!hover { background-color: rgb(69, 90, 100) }
                                            QPushButton:pressed { background-color: black; }
                                        """)
        self.komplektButton.clicked.connect(
            self.make_handleButton("New_Year")
        )

        self.pushButton = QtWidgets.QPushButton("Создать анкету", self.groupBox_2)
        self.pushButton.setGeometry(QtCore.QRect(320, 410, 340, 51))
        self.pushButton.setStyleSheet("""
                            QPushButton:hover { background-color: #bbbbbb }
                            QPushButton:!hover { background-color: rgb(69, 90, 100) }
                            QPushButton:pressed { background-color: black; }
                        """)
        self.pushButton.clicked.connect(
            self.make_handleButton("Create_anketa")
        )

        self.pushButton_2 = QtWidgets.QPushButton("Просмотреть/изменить анкеты", self.groupBox_2)
        self.pushButton_2.setGeometry(QtCore.QRect(320, 490, 340, 51))
        self.pushButton_2.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb }
                                    QPushButton:!hover { background-color: rgb(69, 90, 100) }
                                    QPushButton:pressed { background-color: black; }
                                """)
        self.pushButton_2.clicked.connect(
            self.make_handleButton("Prosmotret_abiturientov")
        )
        self.otchet = QtWidgets.QPushButton("Вывод отчётов", self.groupBox_2)
        self.otchet.setGeometry(QtCore.QRect(320, 570, 340, 51))
        self.otchet.setStyleSheet("""
                                            QPushButton:hover { background-color: #bbbbbb }
                                            QPushButton:!hover { background-color: rgb(69, 90, 100) }
                                            QPushButton:pressed { background-color: black; }
                                        """)
        self.otchet.clicked.connect(
            self.make_handleButton("Otchet")
        )
        self.upload = QtWidgets.QPushButton("Загрузить начальную таблицу", self.groupBox_2)
        self.upload.setGeometry(QtCore.QRect(320, 330, 340, 51))
        self.upload.setStyleSheet("""
                                                   QPushButton:hover { background-color: #bbbbbb }
                                                   QPushButton:!hover { background-color: rgb(69, 90, 100) }
                                                   QPushButton:pressed { background-color: black; }
                                               """)
        self.upload.clicked.connect(self.getFileName)
    def make_handleButton(self, button):
        def handleButton():
            if button == "Create_anketa":
                self.goto("Create_anketa")
            elif button == "Prosmotret_abiturientov":
                self.goto("Prosmotret_abiturientov")
            elif button == "New_Year":
                self.goto("New_Year")
            elif button == "Otchet":
                self.goto("Otchet")
        return handleButton
    def getFileName(self):
        file_name = QFileDialog.getOpenFileName(self, "Выбор таблицы поступивших", None, "Excel File (*.xlsx *.xls)")[0]
        if not file_name:
            return
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Лист1')
        col_names = []
        i = 1
        while sheet.cell(1, i).value is not None:
            col_names.append(sheet.cell(1, i).value)
            i += 1
        max_col = i-1
        print(max_col)
        # Проверка на пустую таблицу
        if max_col < 1:
            exit('Таблица пуста')
        # Найдем столбец с самой большой длиной, из-за возможных пропущенных значений
        j = 1
        max_line = -1
        for n in [1, max_col]:
            while sheet.cell(j, n).value:
                j += 1
            if j > max_line:
                max_line = j-1  # differ 2 because of last j+=1 and also we don't need names line
        print(max_line)
        # Проверка на пустую таблицу
        if max_line < 1:
            exit('Таблица пуста')
        i=1
        self.con()
        while i<=max_line:
            surname = sheet.cell(i,1).value
            name = sheet.cell(i,2).value
            otchestvo = sheet.cell(i,3).value
            zvanie = sheet.cell(i,4).value
            i+=1
            self.cur.execute(
                "INSERT INTO abiturients ( surname_a, name_a, midname_a, zvanie) VALUES ('" + surname + "','" + name + "','" + otchestvo + "','" + zvanie + "')")
            self.conn.commit()
        self.conn.close()
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Запись успешно произведена !")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()


class New_Year(PageWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Создать новый год набора")
        self.UiComponents()

    def goToMain(self):
        self.goto("Abiturient")
    def goToHome(self):
        self.goto("main")
    def UiComponents(self):
        self.setStyleSheet("background-color: white;")
        self.groupBox = QtWidgets.QGroupBox(self)
        self.groupBox.setEnabled(True)
        self.groupBox.setGeometry(QtCore.QRect(-1, 0, 71, 831))
        self.groupBox.setStyleSheet("background-color: rgb(69, 90, 100);")
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.backButton = QtWidgets.QPushButton(self.groupBox)
        self.backButton.setGeometry(QtCore.QRect(100, 5, 100, 20))
        self.backButton.clicked.connect(self.goToMain)
        self.backButton.setGeometry(QtCore.QRect(10, 10, 50, 50))
        self.backButton.setStyleSheet("border-image: url(source/back.png);")
        self.homeButton = QtWidgets.QPushButton(self.groupBox)
        self.homeButton.setGeometry(QtCore.QRect(10, 70, 50, 50))
        self.homeButton.setStyleSheet("border-image: url(source/home.png);")
        self.homeButton.clicked.connect(self.goToHome)
        self.groupBox_2 = QtWidgets.QGroupBox(self)
        self.groupBox_2.setGeometry(QtCore.QRect(100, 20, 1010, 780))
        self.groupBox_2.setStyleSheet("font: 14pt \"Arial\";\n"
                                      "border: 1px solid rgb(69, 90, 100);\n"
                                      "border-radius: 4;\n"
                                      "background-color: rgb(69, 90, 100);\n"
                                      "color: white;")
        self.spec = QtWidgets.QLabel("Специальность", self.groupBox_2)
        self.spec.setGeometry(QtCore.QRect(80, 165, 200, 40))
        self.label = QtWidgets.QLabel(self.groupBox_2)
        self.label.setText("Год набора")
        self.label.setGeometry(QtCore.QRect(80, 215, 200, 40))
        self.label_2 = QtWidgets.QLabel("Срок обучения", self.groupBox_2)
        self.label_2.setGeometry(QtCore.QRect(80, 265, 200, 40))


        self.input = QtWidgets.QLineEdit(self.groupBox_2)
        self.input.setGeometry(QtCore.QRect(300, 215, 550, 40))
        self.input.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")

        self.input_2 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_2.setGeometry(QtCore.QRect(300, 265, 550, 40))
        self.input_2.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        strList_srok = [
            '2 года 5 месяцев', '3 года', '2 года 5 месяцев','3 года','4 года','5 лет 10 месяцев', '5 лет 11 месяцев', '6 лет'
        ]
        self.completer = QtWidgets.QCompleter(strList_srok, self.input_2)
        self.input_2.setCompleter(self.completer)

        # выпадающая специальность
        self.spec = QtWidgets.QComboBox(self.groupBox_2)
        self.spec.addItems(['---',
                            'Специальность 40.02.02 Правоохранительная деятельность',
                            'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Административная деятельность полиции',
                            'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Оперативно-розыскная деятельность полиции',
                            'Специальность 40.05.01 Правовое обеспечение национальной безопасности',
                            'Специальность 44.05.01 Педагогика и психология девиантного поведения',
                            'Направление подготовки 40.04.01 Юриспруденция'])
        self.spec.setGeometry(QtCore.QRect(300, 165, 550, 40))
        self.spec.setStyleSheet("background-color: white;\n"
                                "color: rgb(69, 90, 100);")

        self.pushButton = QtWidgets.QPushButton("Создать год обучения", self.groupBox_2)
        self.pushButton.setGeometry(QtCore.QRect(580, 410, 240, 50))
        self.pushButton.setStyleSheet("""
                            QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                            QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                            QPushButton:pressed { background-color: black; }
                        """)
        self.pushButton.clicked.connect(self.save)
    def save(self):
        self.special = self.spec.currentText()
        self.year = self.input.text()
        self.srok = self.input_2.text()
        self.con()
        print("Database opened successfully")
        self.cur.execute("INSERT INTO kyrses (year_nabor, speciality, srok_obuchenia) VALUES ('" + self.year + "','" + self.special + "','" + self.srok + "')")
        self.conn.commit()
        print("Record inserted successfully")
        self.conn.close()
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Год создан !")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

class Slushatel(PageWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Слушатель")
        self.UiComponents()

    def goToMain(self):
        self.goto("main")
    def goToHome(self):
        self.goto("main")
    def UiComponents(self):
        self.setStyleSheet("background-color: white;")
        self.groupBox = QtWidgets.QGroupBox(self)
        self.groupBox.setEnabled(True)
        self.groupBox.setGeometry(QtCore.QRect(-1, 0, 71, 831))
        self.groupBox.setStyleSheet("background-color: rgb(69, 90, 100);")
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.groupBox_2 = QtWidgets.QGroupBox(self)
        self.groupBox_2.setGeometry(QtCore.QRect(100, 20, 1010, 780))
        self.groupBox_2.setStyleSheet("font: 14pt \"Arial\";\n"
                                      "color: white;\n"
                                      "border: 2px solid rgb(69, 90, 100);\n"
                                      "border-radius: 10;\n"
                                      )
        self.backButton = QtWidgets.QPushButton(self.groupBox)
        self.backButton.setGeometry(QtCore.QRect(100, 5, 100, 20))
        self.backButton.clicked.connect(self.goToMain)
        self.backButton.setGeometry(QtCore.QRect(10, 10, 50, 50))
        self.backButton.setStyleSheet("border-image: url(source/back.png);")
        self.homeButton = QtWidgets.QPushButton(self.groupBox)
        self.homeButton.setGeometry(QtCore.QRect(10, 70, 50, 50))
        self.homeButton.setStyleSheet("border-image: url(source/home.png);")
        self.homeButton.clicked.connect(self.goToHome)
        self.pushButton = QtWidgets.QPushButton("Просмотреть/изменить анкеты", self.groupBox_2)
        self.pushButton.setGeometry(QtCore.QRect(320, 340, 360, 51))
        self.pushButton.setStyleSheet("""
                                            QPushButton:hover { background-color: #bbbbbb }
                                            QPushButton:!hover { background-color: rgb(69, 90, 100) }
                                            QPushButton:pressed { background-color: black; }
                                        """)
        self.pushButton.clicked.connect(
            self.make_handleButton("God_nabora")
        )
        self.pushButton_2 = QtWidgets.QPushButton("Кафедры/дисциплины/преподаватели", self.groupBox_2)
        self.pushButton_2.setGeometry(QtCore.QRect(320, 420, 360, 51))
        self.pushButton_2.setStyleSheet("""
                                            QPushButton:hover { background-color: #bbbbbb }
                                            QPushButton:!hover { background-color: rgb(69, 90, 100) }
                                            QPushButton:pressed { background-color: black; }
                                        """)
        self.pushButton_2.clicked.connect(
            self.make_handleButton("Kafedra")
        )
    def make_handleButton(self, button):
        def handleButton():
            if button == "God_nabora":
                self.msg = QMessageBox()
                self.msg.setWindowTitle("Уведомление")
                self.msg.setText("Модуль находится в стадии разработки !")
                self.msg.setIcon(QMessageBox.Warning)
                self.msg.exec_()
            elif button == "Kafedra":
                self.msg = QMessageBox()
                self.msg.setWindowTitle("Уведомление")
                self.msg.setText("Модуль находится в стадии рзработки!")
                self.msg.setIcon(QMessageBox.Warning)
                self.msg.exec_()

        return handleButton
class Otchet(PageWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Вывести отчёты")
        self.UiComponents()

    def goToMain(self):
        self.goto("Abiturient")

    def goToHome(self):
        self.goto("main")

    def UiComponents(self):
        self.setStyleSheet("background-color: white;")
        self.groupBox = QtWidgets.QGroupBox(self)
        self.groupBox.setEnabled(True)
        self.groupBox.setGeometry(QtCore.QRect(-1, 0, 71, 831))
        self.groupBox.setStyleSheet("background-color: rgb(69, 90, 100);")
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.backButton = QtWidgets.QPushButton(self.groupBox)
        self.backButton.setGeometry(QtCore.QRect(100, 5, 100, 20))
        self.backButton.clicked.connect(self.goToMain)
        self.backButton.setGeometry(QtCore.QRect(10, 10, 50, 50))
        self.backButton.setStyleSheet("border-image: url(source/back.png);")
        self.homeButton = QtWidgets.QPushButton(self.groupBox)
        self.homeButton.setGeometry(QtCore.QRect(10, 70, 50, 50))
        self.homeButton.setStyleSheet("border-image: url(source/home.png);")
        self.homeButton.clicked.connect(self.goToHome)
        self.groupBox_2 = QtWidgets.QGroupBox(self)
        self.groupBox_2.setGeometry(QtCore.QRect(100, 20, 1010, 780))
        self.groupBox_2.setStyleSheet("font: 14pt \"Arial\";\n"
                                      "border: 1px solid rgb(69, 90, 100);\n"
                                      "border-radius: 4;\n"
                                      "background-color: rgb(69, 90, 100);\n"
                                      "color: white;")
        self.pushButton = QtWidgets.QPushButton(
            "Вне конкурса", self.groupBox_2)
        self.pushButton.setGeometry(QtCore.QRect(80, 100, 240, 50))
        self.pushButton.setStyleSheet("""
                            QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                            QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                            QPushButton:pressed { background-color: black; }
                        """)
        self.pushButton.clicked.connect(self.vne_conk)
        self.pushButton_2 = QtWidgets.QPushButton(
            "Возврат дел", self.groupBox_2)
        self.pushButton_2.setGeometry(QtCore.QRect(380, 100, 240, 50))
        self.pushButton_2.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                """)
        self.pushButton_2.clicked.connect(self.vozvrat_dela)
        self.pushButton_3 = QtWidgets.QPushButton(
            "ВопросЛьготы", self.groupBox_2)
        self.pushButton_3.setGeometry(QtCore.QRect(680, 100, 240, 50))
        self.pushButton_3.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                    """)
        self.pushButton_3.clicked.connect(self.vopr_lgoti)
        self.pushButton_4 = QtWidgets.QPushButton("Bыбыли", self.groupBox_2)
        self.pushButton_4.setGeometry(QtCore.QRect(80, 200, 240, 50))
        self.pushButton_4.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                    """)
        self.pushButton_4.clicked.connect(self.outsiders)
        self.pushButton_5 = QtWidgets.QPushButton(
            "Hе приступили к экз", self.groupBox_2)
        self.pushButton_5.setGeometry(QtCore.QRect(380, 200, 240, 50))
        self.pushButton_5.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                    """)
        self.pushButton_5.clicked.connect(self.not_started)
        self.pushButton_6 = QtWidgets.QPushButton("Hе пришли", self.groupBox_2)
        self.pushButton_6.setGeometry(QtCore.QRect(680, 200, 240, 50))
        self.pushButton_6.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                    """)
        self.pushButton_6.clicked.connect(self.neprishli)
        self.pushButton_7 = QtWidgets.QPushButton(
            "Неявка/двойка", self.groupBox_2)
        self.pushButton_7.setGeometry(QtCore.QRect(80, 300, 240, 50))
        self.pushButton_7.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                    """)
        self.pushButton_7.clicked.connect(self.neyavka)
        self.pushButton_8 = QtWidgets.QPushButton("Поступили", self.groupBox_2)
        self.pushButton_8.setGeometry(QtCore.QRect(380, 300, 240, 50))
        self.pushButton_8.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                    """)
        self.pushButton_8.clicked.connect(self.postupili)
        self.pushButton_9 = QtWidgets.QPushButton("Кол-во принятых дел", self.groupBox_2)
        self.pushButton_9.setGeometry(QtCore.QRect(680, 300, 240, 50))
        self.pushButton_9.setStyleSheet("""
                                            QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                            QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                            QPushButton:pressed { background-color: black; }
                                            """)
        self.pushButton_9.clicked.connect(self.summa)
        self.spec = QtWidgets.QComboBox(self.groupBox_2)
        self.spec.addItems(['Выберите специальность или направление подготовки',
                            'Специальность 40.02.02 Правоохранительная деятельность',
                            'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Административная деятельность полиции',
                            'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Оперативно-розыскная деятельность полиции',
                            'Специальность 40.05.01 Правовое обеспечение национальной безопасности',
                            'Специальность 44.05.01 Педагогика и психология девиантного поведени ',
                            'Направление подготовки 40.04.01 Юриспруденция'])
        self.spec.setGeometry(QtCore.QRect(80, 460, 840, 30))
        self.spec.setStyleSheet("background-color: white;\n"
                                "color: rgb(69, 90, 100);")
        self.comp = QtWidgets.QLineEdit(self.groupBox_2)
        self.comp.setGeometry(QtCore.QRect(80, 500, 840, 30))
        self.comp.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")

    def vne_conk(self):
        self.con()
        self.speciality = self.spec.currentText()
        self.c = self.comp.text()
        if self.speciality == 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE note_a='Да' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa DESC;"""
            )
        elif self.speciality != 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and note_a='Да' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa DESC;"""
            )
        elif self.speciality == 'Выберите специальность или направление подготовки' and self.c != '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE complect='{self.c}' and note_a='Да' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa DESC;"""
            )
        else:
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and complect='{self.c}' and note_a='Да' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa DESC;"""
            )

        self.univer_l = []
        self.year_l = []
        self.complect_l = []
        self.zvanie_l = []
        self.fam_l = []
        self.name_l = []
        self.otch_l = []
        self.spec_l = []
        self.fno_l = []

        for self.mean in self.cur.fetchall():
            self.speciality_a = self.mean[1]
            self.sername_a = self.mean[2]
            self.name_a = self.mean[3]
            self.midname_a = self.mean[4]
            self.data_birth_a = self.mean[5]
            self.zvanie = self.mean[6]
            self.complect = self.mean[7]
            self.place_service = self.mean[8]
            self.group_number_a = self.mean[9]
            self.individual = self.mean[10]
            self.note_a = self.mean[11]
            self.document = self.mean[12]
            self.number_registration = self.mean[13]
            self.seria = self.mean[14]
            self.number = self.mean[15]
            self.who_take = self.mean[16]
            self.year_take = self.mean[17]
            self.specialist = self.mean[18]

            self.spec_l.append(self.speciality_a)
            self.univer_l.append(self.who_take)
            self.complect_l.append(self.complect)
            self.zvanie_l.append(self.zvanie)
            self.fam_l.append(self.sername_a)
            self.name_l.append(self.name_a)
            self.otch_l.append(self.midname_a)
            self.fno_l.append(
                f'{self.sername_a} {self.name_a} {self.midname_a}')
            # self.print(f'{self.sername_a} {self.name_a} {self.midname_a}')
            # self.print('=================================================================================================\n')

        def fno_f(self):
            self.number = 0
            self.num_list = []
            for i in range(0, len(self.fno_l)):
                self.number += 1
                self.num_list.append(self.number)
                self.data = f"Заочная {self.spec_l[i]} , на базе {self.univer_l[i]}  \n\t\t\t\t {self.complect_l[i]}\n\t{str(self.number)} {self.zvanie_l[i]} \t {self.fno_l[i]}"
                print(self.data)
                yield self.data

                # tbl_contents = [{'fno': R(fno)} for fno in zip(fno)]

        self.context = {
            # 'spec': R("\n".join(spec_f())),
            # 'complect': R("\n".join(complect_f())),
            'data': R("\n".join(fno_f(self))),
            'vsego': len(self.fam_l)
        }
        self.word_doc = DocxTemplate('docs\вне конкурса.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\вне конкурса_рез.docx')
        self.conn.close()

    def vozvrat_dela(self):
        self.con()
        self.speciality = self.spec.currentText()
        self.c = self.comp.text()
        if self.speciality == 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality != 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality == 'Выберите специальность или направление подготовки' and self.c != '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        else:
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )

        self.univer_l = []
        self.year_l = []
        self.complect_l = []
        self.zvanie_l = []
        self.fam_l = []
        self.name_l = []
        self.otch_l = []
        self.spec_l = []
        self.fno_l = []

        for mean in self.cur.fetchall():
            self.speciality_a = mean[1]
            self.sername_a = mean[2]
            self.name_a = mean[3]
            self.midname_a = mean[4]
            self.data_birth_a = mean[5]
            self.zvanie = mean[6]
            self.complect = mean[7]
            self.place_service = mean[8]
            self.group_number_a = mean[9]
            self.individual = mean[10]
            self.note_a = mean[11]
            self.document = mean[12]
            self.number_registration = mean[13]
            self.seria = mean[14]
            self.number = mean[15]
            self.who_take = mean[16]
            self.year_take = mean[17]
            self.specialist = mean[18]

            self.spec_l.append(self.speciality_a)
            self.univer_l.append(self.who_take)
            self.complect_l.append(self.complect)
            self.zvanie_l.append(self.zvanie)
            self.fam_l.append(self.sername_a)
            self.name_l.append(self.name_a)
            self.otch_l.append(self.midname_a)
            self.fno_l.append(
                f'{self.sername_a} {self.name_a} {self.midname_a}')

        def fno_f(self):
            self.number = 0
            self.num_list = []
            for i in range(0, len(self.fno_l)):
                self.number += 1
                self.num_list.append(self.number)
                self.data = f"{self.complect_l[i]}\n\t{str(self.number)} {self.zvanie_l[i]} \t {self.fno_l[i]}\n"
                yield self.data

        self.context = {
            'spec': self.speciality,
            'data': R("\n".join(fno_f(self))),
            'vsego': len(self.fam_l)
        }

        self.word_doc = DocxTemplate('docs\Возврат дел.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\Возврат дел_рез.docx')
        self.conn.close()

        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def vopr_lgoti(self):
        self.con()
        self.cur.execute(
            f"""SELECT * FROM abiturients WHERE note_a='Да';"""
        )

        self.univer_l = []
        self.year_l = []
        self.complect_l = []
        self.zvanie_l = []
        self.fam_l = []
        self.name_l = []
        self.otch_l = []
        self.spec_l = []
        self.fno_l = []
        self.group_l = []

        for mean in self.cur.fetchall():
            self.speciality_a = mean[1]
            self.sername_a = mean[2]
            self.name_a = mean[3]
            self.midname_a = mean[4]
            self.data_birth_a = mean[5]
            self.zvanie = mean[6]
            self.complect = mean[7]
            self.place_service = mean[8]
            self.group_number_a = mean[9]
            self.individual = mean[10]
            self.note_a = mean[11]
            self.document = mean[12]
            self.number_registration = mean[13]
            self.seria = mean[14]
            self.number = mean[15]
            self.who_take = mean[16]
            self.year_take = mean[17]
            self.specialist = mean[18]

            self.spec_l.append(self.speciality_a)
            self.group_l.append(self.group_number_a)
            self.univer_l.append(self.who_take)
            self.complect_l.append(self.complect)
            self.zvanie_l.append(self.zvanie)
            self.fam_l.append(self.sername_a)
            self.name_l.append(self.name_a)
            self.otch_l.append(self.midname_a)
            self.fno_l.append(
                f'{self.sername_a} {self.name_a} {self.midname_a}')

        def fno_f(self):
            self.number = 0
            self.num_list = []
            for i in range(0, len(self.fno_l)):
                self.number += 1
                self.num_list.append(self.number)
                self.data = f"{str(self.number)}\t {self.group_l[i]}\t {self.zvanie_l[i]} \t {self.fno_l[i]}"
                yield self.data

        self.context = {
            'data': R("\n".join(fno_f(self))),
            'vsego': len(self.fam_l)
        }

        self.word_doc = DocxTemplate('docs\ВопросЛьготы.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\ВопросЛьготы_рез.docx')
        self.conn.close()

        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def outsiders(self):
        # здесь будут баллы
        self.con()
        self.speciality = self.spec.currentText()
        self.c = self.comp.text()
        if self.speciality == 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality != 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality == 'Выберите специальность или направление подготовки' and self.c != '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        else:
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        self.univer_l = []
        self.year_l = []
        self.complect_l = []
        self.zvanie_l = []
        self.fam_l = []
        self.name_l = []
        self.otch_l = []
        self.spec_l = []
        self.fno_l = []
        self.score_l = []

        for mean in self.cur.fetchall():
            self.speciality_a = mean[1]
            self.sername_a = mean[2]
            self.name_a = mean[3]
            self.midname_a = mean[4]
            self.data_birth_a = mean[5]
            self.zvanie = mean[6]
            self.complect = mean[7]
            self.place_service = mean[8]
            self.group_number_a = mean[9]
            self.individual = mean[10]
            self.note_a = mean[11]
            self.document = mean[12]
            self.number_registration = mean[13]
            self.seria = mean[14]
            self.number = mean[15]
            self.who_take = mean[16]
            self.year_take = mean[17]
            self.specialist = mean[18]
            self.rus = mean[21]
            self.soc = mean[22]
            self.his = mean[23]
            self.md = mean[24]
            self.ind = mean[25]
            self.sum = mean[26]

            self.spec_l.append(self.speciality_a)
            self.univer_l.append(self.who_take)
            self.complect_l.append(self.complect)
            self.zvanie_l.append(self.zvanie)
            self.fam_l.append(self.sername_a)
            self.name_l.append(self.name_a)
            self.otch_l.append(self.midname_a)
            self.fno_l.append(f'{self.sername_a} {self.name_a} {self.midname_a}')
            self.score_l.append(f'{self.soc}  {self.rus}  {self.his}  {self.md}  {self.ind}  {self.sum}')

        def fno_f(self):
            self.number = 0
            self.num_list = []
            for i in range(0, len(self.fno_l)):
                self.number += 1
                self.num_list.append(self.number)
                self.data = f"{str(self.number)}  {self.zvanie_l[i]} \t {self.fno_l[i]} \t {self.score_l[i]}"
                yield self.data

        self.context = {
            'data': R("\n".join(fno_f(self))),
            'vsego': len(self.fam_l)
        }

        self.word_doc = DocxTemplate('docs\выбыли.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\выбыли_рез.docx')
        self.conn.close()

        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def not_started(self):
        self.con()
        self.speciality = self.spec.currentText()
        self.c = self.comp.text()
        if self.speciality == 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality != 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality == 'Выберите специальность или направление подготовки' and self.c != '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        else:
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )

        self.univer_l = []
        self.year_l = []
        self.complect_l = []
        self.zvanie_l = []
        self.fam_l = []
        self.name_l = []
        self.otch_l = []
        self.spec_l = []
        self.fno_l = []

        for mean in self.cur.fetchall():
            self.speciality_a = mean[1]
            self.sername_a = mean[2]
            self.name_a = mean[3]
            self.midname_a = mean[4]
            self.data_birth_a = mean[5]
            self.zvanie = mean[6]
            self.complect = mean[7]
            self.place_service = mean[8]
            self.group_number_a = mean[9]
            self.individual = mean[10]
            self.note_a = mean[11]
            self.document = mean[12]
            self.number_registration = mean[13]
            self.seria = mean[14]
            self.number = mean[15]
            self.who_take = mean[16]
            self.year_take = mean[17]
            self.specialist = mean[18]
            self.rus = mean[21]
            self.soc = mean[22]
            self.his = mean[23]
            self.md = mean[24]
            self.ind = mean[25]
            self.sum = mean[26]

            self.spec_l.append(self.speciality_a)
            self.univer_l.append(self.who_take)
            self.complect_l.append(self.complect)
            self.zvanie_l.append(self.zvanie)
            self.fam_l.append(self.sername_a)
            self.name_l.append(self.name_a)
            self.otch_l.append(self.midname_a)
            self.fno_l.append(
                f'{self.sername_a} {self.name_a} {self.midname_a}')

        def fno_f(self):
            self.number = 0
            self.num_list = []
            for i in range(0, len(self.fno_l)):
                self.number += 1
                self.num_list.append(self.number)
                self.data = f"{str(self.number)}  {self.zvanie_l[i]} \t {self.fno_l[i]}"
                yield self.data

        self.context = {
            'spec': self.speciality,
            'data': R("\n".join(fno_f(self))),
            'vsego': len(self.fam_l)
        }

        self.word_doc = DocxTemplate('docs\не приступили к экз.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\не приступили к экз_рез.docx')
        self.conn.close()

        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def neprishli(self):
        self.con()
        self.speciality = self.spec.currentText()
        self.c = self.comp.text()
        if self.speciality == 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality != 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality == 'Выберите специальность или направление подготовки' and self.c != '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        else:
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )

        self.univer_l = []
        self.year_l = []
        self.complect_l = []
        self.zvanie_l = []
        self.fam_l = []
        self.name_l = []
        self.otch_l = []
        self.spec_l = []
        self.fno_l = []
        self.score_l = []

        for mean in self.cur.fetchall():
            self.speciality_a = mean[1]
            self.sername_a = mean[2]
            self.name_a = mean[3]
            self.midname_a = mean[4]
            self.data_birth_a = mean[5]
            self.zvanie = mean[6]
            self.complect = mean[7]
            self.place_service = mean[8]
            self.group_number_a = mean[9]
            self.individual = mean[10]
            self.note_a = mean[11]
            self.document = mean[12]
            self.number_registration = mean[13]
            self.seria = mean[14]
            self.number = mean[15]
            self.who_take = mean[16]
            self.year_take = mean[17]
            self.specialist = mean[18]
            self.rus = mean[21]
            self.soc = mean[22]
            self.his = mean[23]
            self.md = mean[24]
            self.ind = mean[25]
            self.sum = mean[26]

            self.spec_l.append(self.speciality_a)
            self.univer_l.append(self.who_take)
            self.complect_l.append(self.complect)
            self.zvanie_l.append(self.zvanie)
            self.fam_l.append(self.sername_a)
            self.name_l.append(self.name_a)
            self.otch_l.append(self.midname_a)
            self.fno_l.append(
                f'{self.sername_a} {self.name_a} {self.midname_a}')
            self.score_l.append(
                f'{self.rus}  {self.his}  {self.soc}  {self.md}  {self.sum}')

        def fno_f(self):
            self.number = 0
            self.num_list = []
            for i in range(0, len(self.fno_l)):
                self.number += 1
                self.num_list.append(self.number)
                self.data = f"{str(self.number)}  {self.zvanie_l[i]} \t {self.fno_l[i]} \t {self.score_l[i]}"
                yield self.data

        self.context = {
            'data': R("\n".join(fno_f(self))),
            'vsego': len(self.fam_l)
        }

        self.word_doc = DocxTemplate('docs\не пришли.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\не пришли_рез.docx')
        self.conn.close()

        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def neyavka(self):
        self.con()
        self.speciality = self.spec.currentText()
        self.c = self.comp.text()
        if self.speciality == 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality != 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        elif self.speciality == 'Выберите специальность или направление подготовки' and self.c != '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )
        else:
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by summa;"""
            )

        self.univer_l = []
        self.year_l = []
        self.complect_l = []
        self.zvanie_l = []
        self.fam_l = []
        self.name_l = []
        self.otch_l = []
        self.spec_l = []
        self.fno_l = []
        self.score_l = []

        for mean in self.cur.fetchall():
            self.speciality_a = mean[1]
            self.sername_a = mean[2]
            self.name_a = mean[3]
            self.midname_a = mean[4]
            self.data_birth_a = mean[5]
            self.zvanie = mean[6]
            self.complect = mean[7]
            self.place_service = mean[8]
            self.group_number_a = mean[9]
            self.individual = mean[10]
            self.note_a = mean[11]
            self.document = mean[12]
            self.number_registration = mean[13]
            self.seria = mean[14]
            self.number = mean[15]
            self.who_take = mean[16]
            self.year_take = mean[17]
            self.specialist = mean[18]
            self.rus = mean[21]
            self.soc = mean[22]
            self.his = mean[23]
            self.md = mean[24]
            self.ind = mean[25]
            self.sum = mean[26]

            self.spec_l.append(self.speciality_a)
            self.univer_l.append(self.who_take)
            self.complect_l.append(self.complect)
            self.zvanie_l.append(self.zvanie)
            self.fam_l.append(self.sername_a)
            self.name_l.append(self.name_a)
            self.otch_l.append(self.midname_a)
            self.fno_l.append(
                f'{self.sername_a} {self.name_a} {self.midname_a}')
            self.score_l.append(
                f'{self.soc}  {self.rus}  {self.ind}  {self.his}  {self.md}  {self.sum}')

        def fno_f(self):
            self.number = 0
            self.num_list = []
            for i in range(0, len(self.fno_l)):
                self.number += 1
                self.num_list.append(self.number)
                self.data = f"{str(self.number)}  {self.zvanie_l[i]} \t {self.fno_l[i]} \t {self.score_l[i]}"
                yield self.data

        self.context = {
            'complect': self.c,
            'data': R("\n".join(fno_f(self))),
            'vsego': len(self.fam_l)
        }

        self.word_doc = DocxTemplate('docs\не явка или двойка.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\не явка или двойка_рез.docx')
        self.conn.close()

        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def postupili(self):
        self.con()
        self.speciality = self.spec.currentText()
        self.c = self.comp.text()
        if self.speciality == 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE exams.id_abiturient_exam = abiturients.id_abiturient order by note_a, summa DESC;"""
            )
        elif self.speciality != 'Выберите специальность или направление подготовки' and self.c == '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and exams.id_abiturient_exam = abiturients.id_abiturient order by note_a, summa DESC;"""
            )
        elif self.speciality == 'Выберите специальность или направление подготовки' and self.c != '':
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by note_a, summa DESC;"""
            )
        else:
            self.cur.execute(
                f"""SELECT * FROM abiturients, exams WHERE speciality_a='{self.speciality}' and complect='{self.c}' and exams.id_abiturient_exam = abiturients.id_abiturient order by note_a, summa DESC;"""
            )

        self.univer_l = []
        self.year_l = []
        self.complect_l = []
        self.zvanie_l = []
        self.fam_l = []
        self.name_l = []
        self.otch_l = []
        self.spec_l = []
        self.fno_l = []
        self.score_l = []

        for mean in self.cur.fetchall():
            self.speciality_a = mean[1]
            self.sername_a = mean[2]
            self.name_a = mean[3]
            self.midname_a = mean[4]
            self.data_birth_a = mean[5]
            self.zvanie = mean[6]
            self.complect = mean[7]
            self.place_service = mean[8]
            self.group_number_a = mean[9]
            self.individual = mean[10]
            self.note_a = mean[11]
            self.document = mean[12]
            self.number_registration = mean[13]
            self.seria = mean[14]
            self.number = mean[15]
            self.who_take = mean[16]
            self.year_take = mean[17]
            self.specialist = mean[18]
            self.rus = mean[21]
            self.soc = mean[22]
            self.his = mean[23]
            self.md = mean[24]
            self.ind = mean[25]
            self.sum = mean[26]

            self.spec_l.append(self.speciality_a)
            self.univer_l.append(self.who_take)
            self.complect_l.append(self.complect)
            self.zvanie_l.append(self.zvanie)
            self.fam_l.append(self.sername_a)
            self.name_l.append(self.name_a)
            self.otch_l.append(self.midname_a)
            self.fno_l.append(
                f'{self.sername_a} {self.name_a} {self.midname_a}')
            self.score_l.append(
                f'{self.rus}  {self.his}  {self.soc}  {self.md}  {self.ind}  {self.ind}  {self.sum}')

        def fno_f(self):
            self.number = 0
            self.num_list = []
            for i in range(0, len(self.fno_l)):
                self.number += 1
                self.num_list.append(self.number)
                self.data = f"{str(self.number)} \t {self.zvanie_l[i]} \t {self.fno_l[i]}\t{self.score_l[i]}"
                yield self.data

        self.context = {
            'complect': self.c,
            'data': R("\n".join(fno_f(self))),
            'vsego': len(self.fam_l)
        }

        self.word_doc = DocxTemplate('docs\поступили.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\поступили_рез.docx')
        self.conn.close()

    def summa(self):
        self.con()
        self.list_spec = ['Специальность 40.02.02 Правоохранительная деятельность',
                          'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Административная деятельность полиции',
                          'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Оперативно-розыскная деятельность полиции',
                          'Специальность 40.05.01 Правовое обеспечение национальной безопасности',
                          'Специальность 44.05.01 Педагогика и психология девиантного поведения',
                          'Направление подготовки 40.04.01 Юриспруденция']
        self.df_main = pd.DataFrame(
            {'Специальность 40.02.02 Правоохранительная деятельность': pd.Series([], dtype='int'),
             'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Административная деятельность полиции': pd.Series(
                 [], dtype='int'),
             'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Оперативно-розыскная деятельность полиции': pd.Series(
                 [], dtype='int'),
             'Специальность 40.05.01 Правовое обеспечение национальной безопасности': pd.Series([], dtype='int'),
             'Специальность 44.05.01 Педагогика и психология девиантного поведения': pd.Series([], dtype='int'),
             'Направление подготовки 40.04.01 Юриспруденция': pd.Series([], dtype='int')})

        self.cur.execute(
            f"""SELECT DISTINCT complect FROM abiturients;"""
        )
        for i in self.cur.fetchall():
            self.compl = f'{i[0]}'
            self.cur.execute(
                f"""SELECT DISTINCT speciality_a FROM abiturients;"""
            )
            self.num = []
            for j in self.list_spec:
                
                self.spec = j
                self.cur.execute(
                    f"""SELECT count(id_abiturient) FROM abiturients WHERE complect='{self.compl}' and speciality_a='{self.spec}';"""
                )
                self.n = self.cur.fetchone()[0]
                self.num.append(self.n)
                
            self.df_main = pd.concat(
                    [self.df_main, pd.DataFrame({'Специальность 40.02.02 Правоохранительная деятельность': f'{self.num[0]}',
                                             'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Административная деятельность полиции': f'{self.num[1]}',
                                             'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Оперативно-розыскная деятельность полиции': f'{self.num[2]}',
                                             'Специальность 40.05.01 Правовое обеспечение национальной безопасности': f'{self.num[3]}',
                                             'Специальность 44.05.01 Педагогика и психология девиантного поведения': f'{self.num[4]}',
                                             'Направление подготовки 40.04.01 Юриспруденция': f'{self.num[5]}'},
                                            index=[self.compl], columns=self.df_main.columns)])
        self.df_main.to_excel('res/test_table.xlsx')
        self.conn.close()

        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

class Create_anketa(PageWindow):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Создать анкету")
        self.UiComponents()

    def goToHome(self):
        self.goto("main")
    def UiComponents(self):
        self.setStyleSheet("background-color: white;")
        self.groupBox = QtWidgets.QGroupBox(self)
        self.groupBox.setEnabled(True)
        self.groupBox.setGeometry(QtCore.QRect(-1, 0, 71, 831))
        self.groupBox.setStyleSheet("background-color: rgb(69, 90, 100);")
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.backButton = QtWidgets.QPushButton(self.groupBox)
        self.backButton.setGeometry(QtCore.QRect(100, 5, 100, 20))
        self.backButton.clicked.connect(self.goToMain)
        self.backButton.setGeometry(QtCore.QRect(10, 10, 50, 50))
        self.backButton.setStyleSheet("border-image: url(source/back.png);")
        self.homeButton = QtWidgets.QPushButton(self.groupBox)
        self.homeButton.setGeometry(QtCore.QRect(10, 70, 50, 50))
        self.homeButton.setStyleSheet("border-image: url(source/home.png);")
        self.homeButton.clicked.connect(self.goToHome)
        self.groupBox_2 = QtWidgets.QGroupBox(self)
        self.groupBox_2.setGeometry(QtCore.QRect(100, 20, 1010, 780))
        self.groupBox_2.setStyleSheet("font: 12pt \"Arial\";\n"
                                      "border: 1px solid rgb(69, 90, 100);\n"
                                      "border-radius: 4;\n"
                                      "background-color: rgb(69, 90, 100);\n"
                                      "color: white;")
        self.spec = QtWidgets.QLabel("Специальность", self.groupBox_2)
        self.spec.setGeometry(QtCore.QRect(20, 7, 200, 25))
        self.label = QtWidgets.QLabel(self.groupBox_2)
        self.label.setText("Фамилия")
        self.label.setGeometry(QtCore.QRect(20, 40, 200, 25))
        self.label_2 = QtWidgets.QLabel("Имя", self.groupBox_2)
        self.label_2.setGeometry(QtCore.QRect(20, 75, 200, 25))
        self.label_3 = QtWidgets.QLabel("Отчество", self.groupBox_2)
        self.label_3.setGeometry(QtCore.QRect(20, 110, 200, 25))
        self.label_4 = QtWidgets.QLabel("Дата рождения", self.groupBox_2)
        self.label_4.setGeometry(QtCore.QRect(20, 145, 200, 25))
        self.label_5 = QtWidgets.QLabel("Звание", self.groupBox_2)
        self.label_5.setGeometry(QtCore.QRect(20, 180, 200, 25))
        self.label_6 = QtWidgets.QLabel("Комплектующий орган", self.groupBox_2) #Нужен или нет?
        self.label_6.setGeometry(QtCore.QRect(20, 215, 300, 25))
        self.label_tel = QtWidgets.QLabel("Номер телефона", self.groupBox_2)
        self.label_tel.setGeometry(QtCore.QRect(20, 285, 200, 25))
        self.label_12 = QtWidgets.QLabel("Номер группы", self.groupBox_2)
        self.label_12.setGeometry(QtCore.QRect(20, 320, 200, 25))
        self.label_8 = QtWidgets.QLabel("Результаты вступительных испытаний", self.groupBox_2)
        self.label_8.setGeometry(QtCore.QRect(20, 370, 400, 25))
        self.label_9 = QtWidgets.QLabel("Русский язык", self.groupBox_2)
        self.label_9.setGeometry(QtCore.QRect(20, 405, 250, 25))
        self.label_10 = QtWidgets.QLabel("История", self.groupBox_2)
        self.label_10.setGeometry(QtCore.QRect(20, 440, 250, 25))
        self.label_13 = QtWidgets.QLabel("Обществознание", self.groupBox_2)
        self.label_13.setGeometry(QtCore.QRect(20, 475, 250, 25))
        self.label_14 = QtWidgets.QLabel("Междисциплинарный экзамен", self.groupBox_2)
        self.label_14.setGeometry(QtCore.QRect(20, 510, 250, 25))
        self.label_15 = QtWidgets.QLabel("Участие в боевых действиях", self.groupBox_2)
        self.label_15.setGeometry(QtCore.QRect(20, 545, 250, 25))
        self.label_nabor = QtWidgets.QLabel("Год набора", self.groupBox_2)
        self.label_nabor.setGeometry(QtCore.QRect(20, 580, 250, 25))
        self.label_17 = QtWidgets.QLabel("Индивидуальные достижения", self.groupBox_2)
        self.label_17.setGeometry(QtCore.QRect(400, 370, 400, 25))
        self.label_18 = QtWidgets.QLabel("Место службы", self.groupBox_2)
        self.label_18.setGeometry(QtCore.QRect(20, 250, 250, 25))
        self.doc = QtWidgets.QLabel("Документ об образовании", self.groupBox_2)
        self.doc.setGeometry(QtCore.QRect(430, 40, 200, 25))
        self.vuz = QtWidgets.QLabel("Учебное заведение", self.groupBox_2)
        self.vuz.setGeometry(QtCore.QRect(430, 75, 200, 25))
        self.speciality = QtWidgets.QLabel("Специальность", self.groupBox_2)
        self.speciality.setGeometry(QtCore.QRect(430, 110, 200, 25))
        self.data_vidachi = QtWidgets.QLabel("Дата выдачи", self.groupBox_2)
        self.data_vidachi.setGeometry(QtCore.QRect(430, 145, 200, 25))
        self.reg = QtWidgets.QLabel("Регистрационный номер", self.groupBox_2)
        self.reg.setGeometry(QtCore.QRect(430, 180, 200, 25))
        self.seria = QtWidgets.QLabel("Серия диплома", self.groupBox_2)
        self.seria.setGeometry(QtCore.QRect(430, 215, 200, 25))
        self.seria = QtWidgets.QLabel("Номер диплома", self.groupBox_2)
        self.seria.setGeometry(QtCore.QRect(430, 250, 200, 25))

        #окна ввода

        self.doc_input = QtWidgets.QComboBox(self.groupBox_2)
        self.doc_input.addItems(['---',
                            'Аттестат о среднем (полном) общем образовании.',
                            'Диплом о начальном профессиональном образовании.',
                            'Диплом о среднем профессиональном образовании .',
                            'Аттестат о среднем общем образовании.',
                            'Диплом специалиста.',
                            'Диплом бакалавра.',
                            'Диплом магистра.',])
        self.doc_input.setGeometry(QtCore.QRect(640, 40, 350, 25))
        self.doc_input.setStyleSheet("background-color: white;\n"
                                "color: rgb(69, 90, 100);")
        self.vuz_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.vuz_input.setGeometry(QtCore.QRect(640, 75, 350, 25))
        self.vuz_input.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.speciality_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.speciality_input.setGeometry(QtCore.QRect(640, 110, 350, 25))
        self.speciality_input.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.data_vidachi_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.data_vidachi_input.setGeometry(QtCore.QRect(640, 145, 350, 25))
        self.data_vidachi_input.setInputMask("##.##.####")
        self.data_vidachi_input.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.reg_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.reg_input.setGeometry(QtCore.QRect(640, 180, 350, 25))
        self.reg_input.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.seria_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.seria_input.setGeometry(QtCore.QRect(640, 215, 350, 25))
        self.seria_input.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.number_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.number_input.setGeometry(QtCore.QRect(640, 250, 350, 25))
        self.number_input.setStyleSheet("background-color: white;\n"
                                       "color: rgb(69, 90, 100);")


        self.input = QtWidgets.QLineEdit(self.groupBox_2)
        self.input.setGeometry(QtCore.QRect(200, 40, 200, 25))
        self.input.setStyleSheet("background-color: white;\n"
                               "color: rgb(69, 90, 100);")
        self.input_2 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_2.setGeometry(QtCore.QRect(200, 75, 200, 25))
        self.input_2.setStyleSheet("background-color: white;\n"
                               "color: rgb(69, 90, 100);")
        self.input_3 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_3.setGeometry(QtCore.QRect(200, 110, 200, 25))
        self.input_3.setStyleSheet("background-color: white;\n"
                               "color: rgb(69, 90, 100);")
        self.input_4 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_4.setInputMask("##.##.####")
        self.input_4.setGeometry(QtCore.QRect(200, 145, 200, 25))
        self.input_4.setStyleSheet("background-color: white;\n"
                               "color: rgb(69, 90, 100);")
        #выпадающая специальность
        self.spec = QtWidgets.QComboBox(self.groupBox_2)
        self.spec.addItems(['---',
             'Специальность 40.02.02 Правоохранительная деятельность',
             'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Административная деятельность полиции',
             'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Оперативно-розыскная деятельность полиции',
             'Специальность 40.05.01 Правовое обеспечение национальной безопасности',
             'Специальность 44.05.01 Педагогика и психология девиантного поведения',
             'Направление подготовки 40.04.01 Юриспруденция'])
        self.spec.setGeometry(QtCore.QRect(200, 7, 790, 25))
        self.spec.setStyleSheet("background-color: white;\n"
                                         "color: rgb(69, 90, 100);")
        #участие в боевых
        self.boev = QtWidgets.QComboBox(self.groupBox_2)
        self.boev.addItems(['Нет',
                            'Да'])
        self.boev.setGeometry(QtCore.QRect(260, 545, 55, 25))
        self.boev.setStyleSheet("background-color: white;\n"
                                "color: rgb(69, 90, 100);")

        self.input_god_nabora = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_god_nabora.setGeometry(QtCore.QRect(260, 580, 55, 25))
        self.input_god_nabora.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")

        #автозаполнение звания
        self.editor_zvanie = QtWidgets.QLineEdit(self.groupBox_2)
        self.editor_zvanie.setGeometry(QtCore.QRect(200, 180, 200, 25))
        self.editor_zvanie.setStyleSheet("background-color: white;\n"
                                "color: rgb(69, 90, 100);")
        strList_zvania = [
            'без звания','рядовой внутренней службы', 'младший сержант внутренней службы' ,'сержант внутренней службы','сержант внутренней службы','старший сержант внутренней службы',
            'старшина внутренней службы', 'прапорщик внутренней службы', 'старший прапорщик внутренней службы', 'младший лейтенант внутренней службы', 'лейтенант внутренней службы', 'старший лейтенант внутренней службы',
            'капитан внутренней службы', 'майор внутренней службы', 'подполковник внутренней службы', 'полковник внутренней службы',
            'рядовой полиции', 'младший сержант полиции' ,'сержант полиции','сержант полиции','старший сержант полиции',
            'старшина полиции', 'прапорщик полиции', 'старший прапорщик полиции', 'младший лейтенант полиции', 'лейтенант полиции', 'старший лейтенант полиции',
            'капитан полиции', 'майор полиции', 'подполковник полиции', 'полковник полиции',
            'рядовой юстиции', 'младший сержант юстиции', 'сержант юстиции', 'сержант юстиции',
            'старший сержант юстиции',
            'старшина юстиции', 'прапорщик юстиции', 'старший прапорщик юстиции', 'младший лейтенант юстиции',
            'лейтенант юстиции', 'старший лейтенант юстиции',
            'капитан юстиции', 'майор юстиции', 'подполковник юстиции', 'полковник юстиции',
            'рядовой', 'младший сержант', 'сержант', 'сержант',
            'старший сержант',
            'старшина', 'прапорщик', 'старший прапорщик', 'младший лейтенант',
            'лейтенант', 'старший лейтенант',
            'капитан', 'майор', 'подполковник', 'полковник','ефрейтор','мичман'
        ]
        self.completer = QtWidgets.QCompleter(strList_zvania, self.editor_zvanie)
        self.editor_zvanie.setCompleter(self.completer)
        #автозаполнение комплектующего
        self.editor_komplekt = QtWidgets.QLineEdit(self.groupBox_2)
        self.editor_komplekt.setGeometry(QtCore.QRect(200, 215, 200, 25))
        self.editor_komplekt.setStyleSheet("background-color: white;\n"
                                         "color: rgb(69, 90, 100);")
        strList_komplekt = [
            'ГУ МВД России по г. Москве', 'ГУ МВД России по Московской области', 'МВД по Кабардино-Балканской республике',
            'УМВД Росии по Брянской области', 'УМВД Росии по Владимирской области', 'УМВД Росии по Ивановской области','УМВД Росии по Калужской области',
            'УМВД Росии по Смоленской области', 'УМВД Росии по Смоленской области', 'УМВД Росии по Тамбовской области', 'УМВД Росии по Тверской области',
            'УМВД Росии по Тульской области', 'УМВД Росии по Ярославской области', 'БСТМ МВД России', 'ОПБ МВД России', 'ГУВМ МВД России', 'ФКУ "ГИАЦ МВД России"',
            'ФКУ "ГЦАХ и ТО МВД России"','ФКУ НПО "СТиС МВД России"', 'Росгвардия', 'ГФС России', 'УТ МВД России по ЦФО', 'Восточно-Сибирское ЛУ МВД России на транспорте',
            'Московский университет МВД России имени В.Я. Кикотя', 'УТ МВД России по ДФО'
        ]
        self.completer = QtWidgets.QCompleter(strList_komplekt, self.editor_komplekt)
        self.editor_komplekt.setCompleter(self.completer)


        self.input_8 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_8.setGeometry(QtCore.QRect(200, 250, 200, 25))
        self.input_8.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        self.input_tel = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_tel.setGeometry(QtCore.QRect(200, 285, 200, 25))
        self.input_tel.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        self.input_9 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_9.setGeometry(QtCore.QRect(200, 320, 50, 25))
        self.input_9.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        self.ege = QtWidgets.QLineEdit(self.groupBox_2)
        self.ege.setGeometry(QtCore.QRect(260, 405, 50, 25))
        self.ege.setStyleSheet("background-color: white;\n"
                               "color: rgb(69, 90, 100);")
        self.ege_2 = QtWidgets.QLineEdit(self.groupBox_2)
        self.ege_2.setGeometry(QtCore.QRect(260, 440, 50, 25))
        self.ege_2.setStyleSheet("background-color: white;\n"
                               "color: rgb(69, 90, 100);")
        self.ege_3 = QtWidgets.QLineEdit(self.groupBox_2)
        self.ege_3.setGeometry(QtCore.QRect(260, 475, 50, 25))
        self.ege_3.setStyleSheet("background-color: white;\n"
                               "color: rgb(69, 90, 100);")
        self.ege_4 = QtWidgets.QLineEdit(self.groupBox_2)
        self.ege_4.setGeometry(QtCore.QRect(260, 510, 50, 25))
        self.ege_4.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.indiv = QtWidgets.QComboBox(self.groupBox_2)
        self.indiv.addItems(['Нет', '5.2.1 Наличие аттестата о среднем общем образовании с отличием.',
                             '5.2.2 Наличие спортивного разряда или спортивного звания  (не ниже кандидата в мастера спорта).',
                             '5.2.3 Наличие аттестата о среднем общем образовании с итоговыми отметками только «хорошо» и «отлично» ...',
                             '5.2.4 Наличие результатов участия в олимпиадах...',
                             '5.2.5 Наличие серебряного и (или) золотого знака отличия...',
                             '5.2.6 Осуществление волонтерской (добровольческой) деятельности...',
                             '5.2.7 Наличие диплома бакалавра или диплома специалиста с отличием ',
                             '5.2.8 Наличие результатов всероссийского этапа ВСО...'])
        self.indiv.setGeometry(QtCore.QRect(400, 405, 500, 150))
        self.indiv.setStyleSheet("background-color: white;\n"
                                "color: rgb(69, 90, 100);")
        self.pushButton = QtWidgets.QPushButton("Сохранить анкету", self.groupBox_2)
        self.pushButton.setGeometry(QtCore.QRect(770, 700, 160, 50))
        self.pushButton.setStyleSheet("""
                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                    QPushButton:pressed { background-color: black; }
                """ )

        self.pushButton.clicked.connect(self.save)
    def goToMain(self):
        self.goto("Abiturient")
    def save(self):
        self.special = self.spec.currentText()
        self.surname = self.input.text()
        self.name = self.input_2.text()
        self.otchestvo = self.input_3.text()
        self.complect = self.editor_komplekt.text()
        self.zvanie = self.editor_zvanie.text()
        self.date = self.input_4.text()
        self.group  = self.input_9.text()
        self.mesto  = self.input_8.text()
        self.voin = self.boev.currentText()
        self.tel = self.input_tel.text()
        #по документу
        self.document = self.doc_input.currentText()
        self.reg = self.reg_input.text()
        self.seria = self.seria_input.text()
        self.number = self.number_input.text()
        self.who_take = self.vuz_input.text()
        self.year_take = self.data_vidachi_input.text()
        self.specialist = self.speciality_input.text()
        #экзамены
        self.ind = self.indiv.currentText()
        self.rus = int(self.ege.text())
        self.his = int(self.ege_2.text())
        self.soc  = int(self.ege_3.text())
        self.mezh = int(self.ege_4.text())
        if self.ind == "Нет":
            self.ind_points = 0
        elif self.ind == "5.2.1 Наличие аттестата о среднем общем образовании с отличием.":
            self.ind_points = 8
        elif self.ind == "5.2.2 Наличие спортивного разряда или спортивного звания  (не ниже кандидата в мастера спорта).":
            self.ind_points = 6
        elif self.ind == "5.2.3 Наличие аттестата о среднем общем образовании с итоговыми отметками только «хорошо» и «отлично» ...":
            self.ind_points = 5
        elif self.ind == "5.2.4 Наличие результатов участия в олимпиадах...":
            self.ind_points = 2
        elif self.ind == "5.2.5 Наличие серебряного и (или) золотого знака отличия...":
            self.ind_points = 2
        elif self.ind == "5.2.6 Осуществление волонтерской (добровольческой) деятельности...":
            self.ind_points = 2
        elif self.ind == "5.2.7 Наличие диплома бакалавра или диплома специалиста с отличием ":
            self.ind_points = 8
        else:
            self.ind_points = 3
        print (self.ind_points)
        # self.vsego = self.ege_6.text() #поменял, нужна сумма
        self.vsego = self.rus + self.soc + self.his +self.ind_points +self.mezh
        self.ind_points = str(self.ind_points)
        self.vsego = str(self.vsego)
        self.rus = str(self.rus)
        self.soc = str(self.soc)
        self.mezh = str(self.mezh)
        self.his = str(self.his)
        self.con()

        print("Database opened successfully")
        self.cur.execute("INSERT INTO abiturients (speciality_a, surname_a, name_a, midname_a, data_birth_a, zvanie, complect, place_service, group_number_a,  individual, note_a,  document, number_registration, seria, number, who_take, year_take, specialist, phone_number) VALUES ('" + self.special + "','" + self.surname + "','" + self.name + "','" + self.otchestvo + "','"+ self.date + "','" + self.zvanie + "','" + self.complect + "','" + self.mesto + "','" + self.group + "','" + self.ind + "', '"+ self.voin +"', '"+ self.document +"', '"+ self.reg +"', '"+ self.seria +"', '"+ self.number +"', '"+ self.who_take +"', '"+ self.year_take +"', '"+ self.specialist +"', '"+ self.tel +"')")
        self.cur.execute("INSERT INTO exams (russian, social, history, md, individual_d, summa) VALUES ('" + self.rus + "', '" + self.soc + "', '" + self.his + "', '" + self.mezh + "', '" + self.ind_points + "', '" + self.vsego + "')")
        self.conn.commit()
        print("Record inserted successfully")
        self.conn.close()
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Запись успешно произведена !")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()


class Prosmotret_abiturientov(PageWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.con() #подключение базы к таблице
        self.setWindowTitle("Просмотреть/изменить анкеты")
        self.tb = Tb(self)

    def goToHome(self):
        self.goto("main")
    def initUI(self):
        self.UiComponents()

    def UiComponents(self):
        self.setStyleSheet("background-color: white;")
        self.groupBox = QtWidgets.QGroupBox(self)
        self.groupBox.setEnabled(True)
        self.groupBox.setGeometry(QtCore.QRect(-1, 0, 71, 831))
        self.groupBox.setStyleSheet("font: 12pt \"Arial\";\n"
                                      "border: 1px solid rgb(69, 90, 100);\n"
                                      "border-radius: 4;\n"
                                      "background-color: rgb(69, 90, 100);\n"
                                      "color: white;")
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.groupBox_2 = QtWidgets.QGroupBox(self)
        self.groupBox_2.setGeometry(QtCore.QRect(100, 20, 1010, 780))
        self.groupBox_2.setStyleSheet("font: 12pt \"Arial\";\n"
                                      "border: 1px solid rgb(69, 90, 100);\n"
                                      "border-radius: 4;\n"
                                      "background-color: rgb(69, 90, 100);\n"
                                      "color: white;")
        self.backButton = QtWidgets.QPushButton(self.groupBox)
        self.backButton.setGeometry(QtCore.QRect(100, 5, 100, 20))
        self.backButton.clicked.connect(self.goToMain)
        self.backButton.setGeometry(QtCore.QRect(10, 10, 50, 50))
        self.backButton.setStyleSheet("border-image: url(source/back.png);")
        self.homeButton = QtWidgets.QPushButton(self.groupBox)
        self.homeButton.setGeometry(QtCore.QRect(10, 70, 50, 50))
        self.homeButton.setStyleSheet("border-image: url(source/home.png);")
        self.homeButton.clicked.connect(self.goToHome)


        self.open = QtWidgets.QPushButton("К анкете",self.groupBox_2)
        self.open.setStyleSheet("""
                            QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                            QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                            QPushButton:pressed { background-color: black; }
                        """)
        self.open.setGeometry(QtCore.QRect(885, 20, 90, 30))
        self.open.clicked.connect(
            self.make_handleButton('button')
        )

    def make_handleButton(self, button):
        def handleButton():
                self.goto("Anketa_abiturienta")
        return handleButton

    def goToMain(self):
        self.goto("Abiturient")


class Tb(QTableWidget):

    def __init__(self, wg):
        self.wg = wg  # запомнить окно, в котором эта таблица показывается
        super().__init__(wg)
        self.setGeometry(135, 100, 940, 670)
        self.setAlternatingRowColors(True)
        self.setColumnCount(10)
        self.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed)
        self.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.horizontalHeader().setDefaultSectionSize(30)
        self.horizontalHeader().setSectionResizeMode(
            0, QtWidgets.QHeaderView.Fixed)
        self.horizontalHeader().setSectionResizeMode(
            4, QtWidgets.QHeaderView.ResizeToContents)
        self.horizontalHeader().setMinimumSectionSize(20)
        self.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.verticalHeader().setMinimumSectionSize(25)
        self.verticalHeader().hide()
        self.cellClicked.connect(self.click)
        self.updt() # обновить таблицу

# обновление таблицы
    def updt(self):
        self.clear()
        self.setRowCount(0)
        self.setHorizontalHeaderLabels(['ID','Специальность', 'Фамилия','Имя','Отчество','Дата рождения', 'Звание', 'Комплектующий орган', 'Группа', 'Примечание']) # заголовки столцов
        self.wg.cur.execute(
            "SELECT id_abiturient, speciality_a, surname_a, name_a, midname_a, data_birth_a, zvanie, complect, place_service, group_number_a, note_a FROM abiturients order by surname_a")
        rows = self.wg.cur.fetchall()
        i = 0
        for elem in rows:
            self.setRowCount(self.rowCount() + 1)
            j = 0
            for t in elem: # заполняем внутри строки
                self.setItem(i, j, QTableWidgetItem(str(t).strip()))
                j += 1
            i += 1
        self.resizeColumnsToContents()
    def click(self):
        self.clearSelection()
        self.setSelectionMode(QAbstractItemView.MultiSelection)
        row_num = self.currentRow()
        self.selectRow(row_num)

class Anketa_abiturienta (PageWindow):

    def __init__(self):
        super().__init__()
        self.initUI()
    def initUI(self):
        self.setWindowTitle("Анкета абитуриента")
        self.UiComponents()
    def goToHome(self):
        self.goto("main")
    def goToMain(self):
        self.goto("Prosmotret_abiturientov")
    def UiComponents(self):
        self.setStyleSheet("background-color: white;")
        self.groupBox = QtWidgets.QGroupBox(self)
        self.groupBox.setEnabled(True)
        self.groupBox.setGeometry(QtCore.QRect(-1, 0, 71, 831))
        self.groupBox.setStyleSheet("background-color: rgb(69, 90, 100);")
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.backButton = QtWidgets.QPushButton(self.groupBox)
        self.backButton.setGeometry(QtCore.QRect(100, 5, 100, 20))
        self.backButton.clicked.connect(self.goToMain)
        self.backButton.setGeometry(QtCore.QRect(10, 10, 50, 50))
        self.backButton.setStyleSheet("border-image: url(source/back.png);")
        self.homeButton = QtWidgets.QPushButton(self.groupBox)
        self.homeButton.setGeometry(QtCore.QRect(10, 70, 50, 50))
        self.homeButton.setStyleSheet("border-image: url(source/home.png);")
        self.homeButton.clicked.connect(self.goToHome)
        self.groupBox_2 = QtWidgets.QGroupBox(self)
        self.groupBox_2.setGeometry(QtCore.QRect(100, 20, 1010, 780))
        self.groupBox_2.setStyleSheet("font: 12pt \"Arial\";\n"
                                      "border: 1px solid rgb(69, 90, 100);\n"
                                      "border-radius: 4;\n"
                                      "background-color: rgb(69, 90, 100);\n"
                                      "color: white;")
        self.spec = QtWidgets.QLabel("Специальность", self.groupBox_2)
        self.spec.setGeometry(QtCore.QRect(20, 7, 200, 25))
        self.label = QtWidgets.QLabel(self.groupBox_2)
        self.label.setText("Фамилия")
        self.label.setGeometry(QtCore.QRect(20, 40, 200, 25))
        self.label_2 = QtWidgets.QLabel("Имя", self.groupBox_2)
        self.label_2.setGeometry(QtCore.QRect(20, 75, 200, 25))
        self.label_3 = QtWidgets.QLabel("Отчество", self.groupBox_2)
        self.label_3.setGeometry(QtCore.QRect(20, 110, 200, 25))
        self.label_4 = QtWidgets.QLabel("Дата рождения", self.groupBox_2)
        self.label_4.setGeometry(QtCore.QRect(20, 145, 200, 25))
        self.label_5 = QtWidgets.QLabel("Звание", self.groupBox_2)
        self.label_5.setGeometry(QtCore.QRect(20, 180, 200, 25))
        self.label_6 = QtWidgets.QLabel("Комплектующий орган", self.groupBox_2)  # Нужен или нет?
        self.label_6.setGeometry(QtCore.QRect(20, 215, 300, 25))
        self.label_tel = QtWidgets.QLabel("Номер телефона", self.groupBox_2)
        self.label_tel.setGeometry(QtCore.QRect(20, 285, 200, 25))
        self.label_12 = QtWidgets.QLabel("Номер группы", self.groupBox_2)
        self.label_12.setGeometry(QtCore.QRect(20, 320, 200, 25))
        self.label_8 = QtWidgets.QLabel("Результаты вступительных испытаний", self.groupBox_2)
        self.label_8.setGeometry(QtCore.QRect(20, 370, 400, 25))
        self.label_9 = QtWidgets.QLabel("Русский язык", self.groupBox_2)
        self.label_9.setGeometry(QtCore.QRect(20, 405, 250, 25))
        self.label_10 = QtWidgets.QLabel("История", self.groupBox_2)
        self.label_10.setGeometry(QtCore.QRect(20, 440, 250, 25))
        self.label_13 = QtWidgets.QLabel("Обществознание", self.groupBox_2)
        self.label_13.setGeometry(QtCore.QRect(20, 475, 250, 25))
        self.label_14 = QtWidgets.QLabel("Междисциплинарный экзамен", self.groupBox_2)
        self.label_14.setGeometry(QtCore.QRect(20, 510, 250, 25))
        self.label_15 = QtWidgets.QLabel("Участие в боевых действиях", self.groupBox_2)
        self.label_15.setGeometry(QtCore.QRect(20, 545, 250, 25))
        self.label_16 = QtWidgets.QLabel("Общее кол-во баллов", self.groupBox_2)
        self.label_16.setGeometry(QtCore.QRect(20, 580, 250, 25))
        self.label_17 = QtWidgets.QLabel("Индивидуальные достижения", self.groupBox_2)
        self.label_17.setGeometry(QtCore.QRect(400, 370, 400, 25))
        self.label_18 = QtWidgets.QLabel("Место службы", self.groupBox_2)
        self.label_18.setGeometry(QtCore.QRect(20, 250, 250, 25))
        self.doc = QtWidgets.QLabel("Документ об образовании", self.groupBox_2)
        self.doc.setGeometry(QtCore.QRect(430, 40, 200, 25))
        self.vuz = QtWidgets.QLabel("Учебное заведение", self.groupBox_2)
        self.vuz.setGeometry(QtCore.QRect(430, 75, 200, 25))
        self.speciality = QtWidgets.QLabel("Специальность", self.groupBox_2)
        self.speciality.setGeometry(QtCore.QRect(430, 110, 200, 25))
        self.data_vidachi = QtWidgets.QLabel("Дата выдачи", self.groupBox_2)
        self.data_vidachi.setGeometry(QtCore.QRect(430, 145, 200, 25))
        self.reg = QtWidgets.QLabel("Регистрационный номер", self.groupBox_2)
        self.reg.setGeometry(QtCore.QRect(430, 180, 200, 25))
        self.seria = QtWidgets.QLabel("Серия диплома", self.groupBox_2)
        self.seria.setGeometry(QtCore.QRect(430, 215, 200, 25))
        self.seria = QtWidgets.QLabel("Номер диплома", self.groupBox_2)
        self.seria.setGeometry(QtCore.QRect(430, 250, 200, 25))
        self.god_nabora = QtWidgets.QLabel("Год набора", self.groupBox_2)
        self.god_nabora.setGeometry(QtCore.QRect(20, 615, 200, 25))
        # окна ввода
        self.input_god_nabora = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_god_nabora.setGeometry(QtCore.QRect(260, 615, 55, 25))
        self.input_god_nabora.setStyleSheet("background-color: white;\n"
                                            "color: rgb(69, 90, 100);")
        self.doc_input = QtWidgets.QComboBox(self.groupBox_2)
        self.doc_input.addItems(['---',
                                 'Аттестат о среднем (полном) общем образовании.',
                                 'Диплом о начальном профессиональном образовании.',
                                 'Диплом о среднем профессиональном образовании .',
                                 'Аттестат о среднем общем образовании.',
                                 'Диплом специалиста.',
                                 'Диплом бакалавра.',
                                 'Диплом магистра.', ])
        self.doc_input.setGeometry(QtCore.QRect(640, 40, 350, 25))
        self.doc_input.setStyleSheet("background-color: white;\n"
                                     "color: rgb(69, 90, 100);")
        self.vuz_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.vuz_input.setGeometry(QtCore.QRect(640, 75, 350, 25))
        self.vuz_input.setStyleSheet("background-color: white;\n"
                                     "color: rgb(69, 90, 100);")
        self.speciality_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.speciality_input.setGeometry(QtCore.QRect(640, 110, 350, 25))
        self.speciality_input.setStyleSheet("background-color: white;\n"
                                            "color: rgb(69, 90, 100);")
        self.data_vidachi_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.data_vidachi_input.setGeometry(QtCore.QRect(640, 145, 350, 25))
        self.data_vidachi_input.setInputMask("##.##.####")
        self.data_vidachi_input.setStyleSheet("background-color: white;\n"
                                              "color: rgb(69, 90, 100);")
        self.reg_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.reg_input.setGeometry(QtCore.QRect(640, 180, 350, 25))
        self.reg_input.setStyleSheet("background-color: white;\n"
                                     "color: rgb(69, 90, 100);")
        self.seria_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.seria_input.setGeometry(QtCore.QRect(640, 215, 350, 25))
        self.seria_input.setStyleSheet("background-color: white;\n"
                                       "color: rgb(69, 90, 100);")
        self.number_input = QtWidgets.QLineEdit(self.groupBox_2)
        self.number_input.setGeometry(QtCore.QRect(640, 250, 350, 25))
        self.number_input.setStyleSheet("background-color: white;\n"
                                        "color: rgb(69, 90, 100);")

        self.input = QtWidgets.QLineEdit(self.groupBox_2)
        self.input.setGeometry(QtCore.QRect(200, 40, 200, 25))
        self.input.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.input_2 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_2.setGeometry(QtCore.QRect(200, 75, 200, 25))
        self.input_2.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        self.input_3 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_3.setGeometry(QtCore.QRect(200, 110, 200, 25))
        self.input_3.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        self.input_4 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_4.setInputMask("##.##.####")
        self.input_4.setGeometry(QtCore.QRect(200, 145, 200, 25))
        self.input_4.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        # выпадающая специальность
        self.spec = QtWidgets.QComboBox(self.groupBox_2)
        self.spec.addItems(['---',
                            'Специальность 40.02.02 Правоохранительная деятельность',
                            'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Административная деятельность полиции',
                            'Направление подготовки 40.03.02 Обеспечение законности и правопорядка. Специализация: Оперативно-розыскная деятельность полиции',
                            'Специальность 40.05.01 Правовое обеспечение национальной безопасности',
                            'Специальность 44.05.01 Педагогика и психология девиантного поведения',
                            'Направление подготовки 40.04.01 Юриспруденция'])
        self.spec.setGeometry(QtCore.QRect(200, 7, 790, 25))
        self.spec.setStyleSheet("background-color: white;\n"
                                "color: rgb(69, 90, 100);")
        # участие в боевых
        self.boev = QtWidgets.QComboBox(self.groupBox_2)
        self.boev.addItems(['Нет',
                            'Да'])
        self.boev.setGeometry(QtCore.QRect(260, 545, 55, 25))
        self.boev.setStyleSheet("background-color: white;\n"
                                "color: rgb(69, 90, 100);")
        # автозаполнение звания
        self.editor_zvanie = QtWidgets.QLineEdit(self.groupBox_2)
        self.editor_zvanie.setGeometry(QtCore.QRect(200, 180, 200, 25))
        self.editor_zvanie.setStyleSheet("background-color: white;\n"
                                         "color: rgb(69, 90, 100);")
        strList_zvania = [
            'без звания', 'рядовой внутренней службы', 'младший сержант внутренней службы', 'сержант внутренней службы',
            'сержант внутренней службы', 'старший сержант внутренней службы',
            'старшина внутренней службы', 'прапорщик внутренней службы', 'старший прапорщик внутренней службы',
            'младший лейтенант внутренней службы', 'лейтенант внутренней службы', 'старший лейтенант внутренней службы',
            'капитан внутренней службы', 'майор внутренней службы', 'подполковник внутренней службы',
            'полковник внутренней службы',
            'рядовой полиции', 'младший сержант полиции', 'сержант полиции', 'сержант полиции',
            'старший сержант полиции',
            'старшина полиции', 'прапорщик полиции', 'старший прапорщик полиции', 'младший лейтенант полиции',
            'лейтенант полиции', 'старший лейтенант полиции',
            'капитан полиции', 'майор полиции', 'подполковник полиции', 'полковник полиции',
            'рядовой юстиции', 'младший сержант юстиции', 'сержант юстиции', 'сержант юстиции',
            'старший сержант юстиции',
            'старшина юстиции', 'прапорщик юстиции', 'старший прапорщик юстиции', 'младший лейтенант юстиции',
            'лейтенант юстиции', 'старший лейтенант юстиции',
            'капитан юстиции', 'майор юстиции', 'подполковник юстиции', 'полковник юстиции',
            'рядовой', 'младший сержант', 'сержант', 'сержант',
            'старший сержант',
            'старшина', 'прапорщик', 'старший прапорщик', 'младший лейтенант',
            'лейтенант', 'старший лейтенант',
            'капитан', 'майор', 'подполковник', 'полковник', 'ефрейтор', 'мичман'
        ]
        self.completer = QtWidgets.QCompleter(strList_zvania, self.editor_zvanie)
        self.editor_zvanie.setCompleter(self.completer)
        # автозаполнение комплектующего
        self.editor_komplekt = QtWidgets.QLineEdit(self.groupBox_2)
        self.editor_komplekt.setGeometry(QtCore.QRect(200, 215, 200, 25))
        self.editor_komplekt.setStyleSheet("background-color: white;\n"
                                           "color: rgb(69, 90, 100);")
        strList_komplekt = [
            'ГУ МВД России по г. Москве', 'ГУ МВД России по Московской области',
            'МВД по Кабардино-Балканской республике',
            'УМВД Росии по Брянской области', 'УМВД Росии по Владимирской области', 'УМВД Росии по Ивановской области',
            'УМВД Росии по Калужской области',
            'УМВД Росии по Смоленской области', 'УМВД Росии по Смоленской области', 'УМВД Росии по Тамбовской области',
            'УМВД Росии по Тверской области',
            'УМВД Росии по Тульской области', 'УМВД Росии по Ярославской области', 'БСТМ МВД России', 'ОПБ МВД России',
            'ГУВМ МВД России', 'ФКУ "ГИАЦ МВД России"',
            'ФКУ "ГЦАХ и ТО МВД России"', 'ФКУ НПО "СТиС МВД России"', 'Росгвардия', 'ГФС России',
            'УТ МВД России по ЦФО', 'Восточно-Сибирское ЛУ МВД России на транспорте',
            'Московский университет МВД России имени В.Я. Кикотя', 'УТ МВД России по ДФО'
        ]
        self.completer = QtWidgets.QCompleter(strList_komplekt, self.editor_komplekt)
        self.editor_komplekt.setCompleter(self.completer)

        self.input_8 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_8.setGeometry(QtCore.QRect(200, 250, 200, 25))
        self.input_8.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        self.input_tel = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_tel.setGeometry(QtCore.QRect(200, 285, 200, 25))
        self.input_tel.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        self.input_9 = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_9.setGeometry(QtCore.QRect(200, 320, 50, 25))
        self.input_9.setStyleSheet("background-color: white;\n"
                                   "color: rgb(69, 90, 100);")
        self.ege = QtWidgets.QLineEdit(self.groupBox_2)
        self.ege.setGeometry(QtCore.QRect(260, 405, 50, 25))
        self.ege.setStyleSheet("background-color: white;\n"
                               "color: rgb(69, 90, 100);")
        self.ege_2 = QtWidgets.QLineEdit(self.groupBox_2)
        self.ege_2.setGeometry(QtCore.QRect(260, 440, 50, 25))
        self.ege_2.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.ege_3 = QtWidgets.QLineEdit(self.groupBox_2)
        self.ege_3.setGeometry(QtCore.QRect(260, 475, 50, 25))
        self.ege_3.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.ege_4 = QtWidgets.QLineEdit(self.groupBox_2)
        self.ege_4.setGeometry(QtCore.QRect(260, 510, 50, 25))
        self.ege_4.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.ege_6 = QtWidgets.QLineEdit(self.groupBox_2)
        self.ege_6.setGeometry(QtCore.QRect(260, 580, 50, 25))
        self.ege_6.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.indiv = QtWidgets.QComboBox(self.groupBox_2)
        self.indiv.addItems(['Нет', '5.2.1 Наличие аттестата о среднем общем образовании с отличием.',
                             '5.2.2 Наличие спортивного разряда или спортивного звания  (не ниже кандидата в мастера спорта).',
                             '5.2.3 Наличие аттестата о среднем общем образовании с итоговыми отметками только «хорошо» и «отлично» ...',
                             '5.2.4 Наличие результатов участия в олимпиадах...',
                             '5.2.5 Наличие серебряного и (или) золотого знака отличия...',
                             '5.2.6 Осуществление волонтерской (добровольческой) деятельности...',
                             '5.2.7 Наличие диплома бакалавра или диплома специалиста с отличием ',
                             '5.2.8 Наличие результатов всероссийского этапа ВСО...'])
        self.indiv.setGeometry(QtCore.QRect(400, 405, 500, 150))
        self.indiv.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.pushButton = QtWidgets.QPushButton("Сохранить анкету", self.groupBox_2)
        self.pushButton.setGeometry(QtCore.QRect(770, 700, 160, 50))
        self.pushButton.setStyleSheet("""
                            QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                            QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                            QPushButton:pressed { background-color: black; }
                        """)

        self.pushButton.clicked.connect(self.save)
        self.label_id = QtWidgets.QLabel("Введите id", self.groupBox_2)
        self.label_id.setGeometry(QtCore.QRect(200, 650, 200, 25))
        self.input_id = QtWidgets.QLineEdit(self.groupBox_2)
        self.input_id.setGeometry(QtCore.QRect(300, 650, 50, 25))
        self.input_id.setStyleSheet("background-color: white;\n"
                                 "color: rgb(69, 90, 100);")
        self.idButton = QtWidgets.QPushButton("Поиск", self.groupBox_2)
        self.idButton.setGeometry(QtCore.QRect(350, 650, 70, 25))
        self.idButton.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                """)
        self.idButton.clicked.connect(self.zapolnenie)
        self.Button_2 = QtWidgets.QPushButton("Переат", self.groupBox_2)
        self.Button_2.setGeometry(QtCore.QRect(20, 700, 120, 50))
        self.Button_2.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                """)
        self.Button_2.clicked.connect(self.pereatt)
        self.Button_3 = QtWidgets.QPushButton("Перезачёт", self.groupBox_2)
        self.Button_3.setGeometry(QtCore.QRect(150, 700, 120, 50))
        self.Button_3.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                """)
        self.Button_3.clicked.connect(self.perezach)
        self.Button_4 = QtWidgets.QPushButton("ВБ АД", self.groupBox_2)
        self.Button_4.setGeometry(QtCore.QRect(280, 700, 120, 50))
        self.Button_4.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                """)
        self.Button_4.clicked.connect(self.vb_ad)
        self.Button_5 = QtWidgets.QPushButton("Индив-й план", self.groupBox_2)
        self.Button_5.setGeometry(QtCore.QRect(410, 700, 120, 50))
        self.Button_5.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                """)
        self.Button_5.clicked.connect(self.ind_plan)
        self.Button_6 = QtWidgets.QPushButton("Справка-вызов", self.groupBox_2)
        self.Button_6.setGeometry(QtCore.QRect(540, 700, 120, 50))
        self.Button_6.setStyleSheet("""
                                    QPushButton:hover { background-color: #bbbbbb; color: rgb(69, 90, 100) }
                                    QPushButton:!hover { background-color: white; color: rgb(69, 90, 100)}
                                    QPushButton:pressed { background-color: black; }
                                """)
        self.Button_6.clicked.connect(self.spravka_vizov)


    def zapolnenie(self):
        self.id_ab = self.input_id.text()
        self.con()
        self.cur.execute(
            "SELECT id_abiturient FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.id_from_pg = str(self.cur.fetchone())[1:-2]
        print(self.id_from_pg)
        self.cur.execute(
            "SELECT surname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.surname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT name_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.name = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT midname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.midname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT data_birth_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data_birth = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT zvanie FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.zvanie_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT complect FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.complect_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT place_service FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.place_service = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT group_number_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.group_number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT russian FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.rus = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT social FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.soc = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT history FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.his = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT summa FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.summa = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT md FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.md = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT number_registration FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number_registration = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT year_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT specialist FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT who_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.vuz = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT seria FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.seria = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT number FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT speciality_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT individual FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.indiv_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT note_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.boev_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT document FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.doc = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT phone_number FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.tel = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT year_nabor  FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.year_nabor = str(self.cur.fetchone())[2:-3]

        self.conn.close()
        #собственно заполняем поля
        self.input_god_nabora.setText(self.year_nabor)
        self.input.setText(self.surname)
        self.input_2.setText(self.name)
        self.input_3.setText(self.midname)
        self.editor_komplekt.setText(self.complect_a)#
        self.editor_zvanie.setText(self.zvanie_a)
        self.input_9.setText(self.group_number)
        self.input_8.setText(self.place_service)#
        self.ege.setText(self.rus)
        self.ege_2.setText(self.his)
        self.ege_3.setText(self.soc)
        self.ege_6.setText(self.summa)
        self.input_4.setText(self.data_birth)
        self.ege_4.setText(self.md)
        self.input_tel.setText(self.tel)

        self.reg_input.setText(self.number_registration)
        self.data_vidachi_input.setText(self.data)
        self.speciality_input.setText(self.speciality)
        self.vuz_input.setText(self.vuz)
        self.seria_input.setText(self.seria)
        self.number_input.setText(self.number)

        self.spec.setCurrentText(self.speciality_a)
        self.indiv.setCurrentText(self.indiv_a)
        self.boev.setCurrentText(self.boev_a)
        self.doc_input.setCurrentText(self.doc)


    def save(self):
        self.special = self.spec.currentText()
        self.surname = self.input.text()
        self.name = self.input_2.text()
        self.otchestvo = self.input_3.text()
        self.complect = self.editor_komplekt.text()
        self.zvanie = self.editor_zvanie.text()
        self.date = self.input_4.text()
        self.group = self.input_9.text()
        self.mesto = self.input_8.text()
        self.voin = self.boev.currentText()
        self.tel = self.input_tel.text()
        self.year_nabor = self.input_god_nabora.text()
        # по документу
        self.document = self.doc_input.currentText()
        self.reg = self.reg_input.text()
        self.seria = self.seria_input.text()
        self.number = self.number_input.text()
        self.who_take = self.vuz_input.text()
        self.year_take = self.data_vidachi_input.text()
        self.specialist = self.speciality_input.text()
        # экзамены
        self.ind = self.indiv.currentText()
        self.rus = int(self.ege.text())
        self.his = int(self.ege_2.text())
        self.soc = int(self.ege_3.text())
        self.mezh = int(self.ege_4.text())
        if self.ind == "Нет":
            self.ind_points = 0
        elif self.ind == "5.2.1 Наличие аттестата о среднем общем образовании с отличием.":
            self.ind_points = 8
        elif self.ind == "5.2.2 Наличие спортивного разряда или спортивного звания  (не ниже кандидата в мастера спорта).":
            self.ind_points = 6
        elif self.ind == "5.2.3 Наличие аттестата о среднем общем образовании с итоговыми отметками только «хорошо» и «отлично» ...":
            self.ind_points = 5
        elif self.ind == "5.2.4 Наличие результатов участия в олимпиадах...":
            self.ind_points = 2
        elif self.ind == "5.2.5 Наличие серебряного и (или) золотого знака отличия...":
            self.ind_points = 2
        elif self.ind == "5.2.6 Осуществление волонтерской (добровольческой) деятельности...":
            self.ind_points = 2
        elif self.ind == "5.2.7 Наличие диплома бакалавра или диплома специалиста с отличием ":
            self.ind_points = 8
        else:
            self.ind_points = 3
        print(self.ind_points)
        self.vsego = self.rus + self.soc + self.his + self.ind_points + self.mezh
        self.ind_points = str(self.ind_points)
        self.vsego = str(self.vsego)
        self.rus = str(self.rus)
        self.soc = str(self.soc)
        self.mezh = str(self.mezh)
        self.his = str(self.his)
        self.ind_points = str(self.ind_points)

        self.con()

        print("Database opened successfully")
        self.cur.execute(
            "DELETE from exams where id_exam = '"+ self.id_from_pg +"'")
        self.cur.execute(
            "DELETE from abiturients where id_abiturient = '" + self.id_from_pg + "'")
        self.cur.execute(
            "INSERT INTO abiturients (speciality_a, surname_a, name_a, midname_a, data_birth_a, zvanie, complect, place_service, group_number_a,  individual, note_a,  document, number_registration, seria, number, who_take, year_take, specialist, phone_number,year_nabor) VALUES ('" + self.special + "','" + self.surname + "','" + self.name + "','" + self.otchestvo + "','" + self.date + "','" + self.zvanie + "','" + self.complect + "','" + self.mesto + "','" + self.group + "','" + self.ind + "', '" + self.voin + "', '" + self.document + "', '" + self.reg + "', '" + self.seria + "', '" + self.number + "', '" + self.who_take + "', '" + self.year_take + "', '" + self.specialist + "', '" + self.tel + "', '"+ self.year_nabora + "')")
        self.cur.execute(
            "INSERT INTO exams (russian, social, history, md, individual_d, summa) VALUES ('" + self.rus + "', '" + self.soc + "', '" + self.his + "', '" + self.mezh + "', '" + self.ind_points + "', '" + self.vsego + "')")
        self.conn.commit()
        print("Record inserted successfully")
        self.conn.close()
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Запись успешно произведена !")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

        # =============================================================================
        # =============================================================================
        # =============================================================================

    def pereatt(self):
        self.con()
        self.id_ab = self.input_id.text()
        self.cur.execute(
            "SELECT id_abiturient FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.id_from_pg = str(self.cur.fetchone())[1:-2]

        print(self.id_from_pg)
        self.cur.execute(
            "SELECT surname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.surname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT name_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.name = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT midname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.midname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT data_birth_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data_birth = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT zvanie FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.zvanie_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT complect FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.complect_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT place_service FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.place_service = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT group_number_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.group_number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT russian FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.rus = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT social FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.soc = str(self.cur.fetchone())[2:-3]
        self.cur.execute(

            "SELECT history FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.his = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT summa FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.summa = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT md FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.md = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT number_registration FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number_registration = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT year_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT specialist FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT who_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.vuz = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT seria FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.seria = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT number FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT speciality_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT individual FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.indiv_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT note_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.boev_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT document FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.doc = str(self.cur.fetchone())[2:-3]

        self.fio = f"{self.surname} {self.name}. {self.midname}."
        self.context = {
            'document': self.doc.replace('диплом', 'диплома'),
            'universitet': self.vuz,
            'num': self.number_registration,
            'year': self.data,
            'fam': self.surname,
            'name': self.name,
            'otch': self.midname,
            'spec': self.speciality_a,
            'fio': self.fio
        }

        self.word_doc = DocxTemplate('docs\ПереаттАД.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\ПереаттАД_рез.docx')
        self.conn.close()
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def perezach(self):
        self.con()
        self.id_ab = self.input_id.text()
        self.cur.execute(
            "SELECT id_abiturient FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.id_from_pg = str(self.cur.fetchone())[1:-2]

        print(self.id_from_pg)
        self.cur.execute(
            "SELECT surname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.surname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT name_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.name = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT midname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.midname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT data_birth_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data_birth = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT zvanie FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.zvanie_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT complect FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.complect_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT place_service FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.place_service = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT group_number_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.group_number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT russian FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.rus = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT social FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.soc = str(self.cur.fetchone())[2:-3]
        self.cur.execute(

            "SELECT history FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.his = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT summa FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.summa = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT md FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.md = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT number_registration FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number_registration = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT year_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT specialist FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT who_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.vuz = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT seria FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.seria = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT number FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT speciality_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT individual FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.indiv_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT note_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.boev_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT document FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.doc = str(self.cur.fetchone())[2:-3]

        self.fio = f"{self.surname} {self.name}. {self.midname}."
        self.context = {
            'document': self.doc.replace('диплом', 'диплома'),
            'universitet': self.vuz,
            'num': self.number_registration,
            'year': self.data,
            'fam': self.surname,
            'name': self.name,
            'otch': self.midname,
            'spec': self.speciality_a,
            'fio': self.fio
        }

        self.word_doc = DocxTemplate('docs\ПерезачетАД2021.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\ПерезачетАД2021_рез.docx')
        self.conn.close()
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def vb_ad(self):
        self.con()
        self.id_ab = self.input_id.text()
        self.cur.execute(
            "SELECT id_abiturient FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.id_from_pg = str(self.cur.fetchone())[1:-2]

        print(self.id_from_pg)
        self.cur.execute(
            "SELECT surname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.surname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT name_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.name = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT midname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.midname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT data_birth_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data_birth = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT zvanie FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.zvanie_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT complect FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.complect_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT place_service FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.place_service = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT group_number_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.group_number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT russian FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.rus = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT social FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.soc = str(self.cur.fetchone())[2:-3]
        self.cur.execute(

            "SELECT history FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.his = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT summa FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.summa = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT md FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.md = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT number_registration FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number_registration = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT year_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT specialist FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT who_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.vuz = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT seria FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.seria = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT number FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT speciality_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT individual FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.indiv_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT note_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.boev_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT document FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.doc = str(self.cur.fetchone())[2:-3]

        self.fio = f"{self.surname} {self.name}. {self.midname}."
        self.context = {
            'document': self.doc.replace('диплом', 'диплома'),
            'universitet': self.vuz,
            'num': self.number_registration,
            'year': self.data,
            'fam': self.surname,
            'name': self.name,
            'otch': self.midname,
            'spec': self.speciality_a,
            'spec1': self.speciality,
            'fio': self.fio
        }

        self.word_doc = DocxTemplate('docs\ВБ_АД2021.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\ВБ_АД2021_рез.docx')
        self.conn.close()
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def ind_plan(self):
        self.con()
        self.id_ab = self.input_id.text()
        self.cur.execute(
            "SELECT id_abiturient FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.id_from_pg = str(self.cur.fetchone())[1:-2]

        print(self.id_from_pg)
        self.cur.execute(
            "SELECT surname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.surname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT name_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.name = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT midname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.midname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT data_birth_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data_birth = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT zvanie FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.zvanie_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT complect FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.complect_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT place_service FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.place_service = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT group_number_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.group_number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT russian FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.rus = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT social FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.soc = str(self.cur.fetchone())[2:-3]
        self.cur.execute(

            "SELECT history FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.his = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT summa FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.summa = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT md FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.md = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT number_registration FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number_registration = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT year_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT specialist FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT who_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.vuz = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT seria FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.seria = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT number FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT speciality_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT individual FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.indiv_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT note_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.boev_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT document FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.doc = str(self.cur.fetchone())[2:-3]

        self.fio = f"{self.surname} {self.name}. {self.midname}."
        self.context = {
            'document': self.doc.replace('диплом', 'диплома'),
            'universitet': self.vuz,
            'num': self.number_registration,
            'year': self.data,
            'fam': self.surname,
            'name': self.name,
            'otch': self.midname,
            'spec': self.speciality_a,
            'fio': self.fio
        }

        self.word_doc = DocxTemplate('docs\Индивидуальный план.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\Индивидуальный план_рез.docx')
        self.conn.close()
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()

    def spravka_vizov(self):
        self.con()
        self.id_ab = self.input_id.text()
        self.cur.execute(
            "SELECT id_abiturient FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.id_from_pg = str(self.cur.fetchone())[1:-2]

        print(self.id_from_pg)
        self.cur.execute(
            "SELECT surname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.surname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT name_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.name = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT midname_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.midname = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT data_birth_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data_birth = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT zvanie FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.zvanie_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT complect FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.complect_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT place_service FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.place_service = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT group_number_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.group_number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT russian FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.rus = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT social FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.soc = str(self.cur.fetchone())[2:-3]
        self.cur.execute(

            "SELECT history FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.his = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT summa FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.summa = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT md FROM exams WHERE id_exam = '" + self.id_ab + "'"
        )
        self.md = str(self.cur.fetchone())[2:-3]

        self.cur.execute(
            "SELECT number_registration FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number_registration = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT year_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.data = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT specialist FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT who_take FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.vuz = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT seria FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.seria = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT number FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.number = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT speciality_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.speciality_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT individual FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.indiv_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT note_a FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.boev_a = str(self.cur.fetchone())[2:-3]
        self.cur.execute(
            "SELECT document FROM abiturients WHERE id_abiturient = '" + self.id_ab + "'"
        )
        self.doc = str(self.cur.fetchone())[2:-3]

        self.morph = pymorphy2.MorphAnalyzer()
        self.fam_w = self.morph.parse(self.surname)[0]
        self.fam_d = self.fam_w.inflect({'sing', 'datv'}).word.capitalize()
        self.name_w = self.morph.parse(self.name)[0]
        self.name_d = self.name_w.inflect({'sing', 'datv'}).word.capitalize()
        self.otch_w = self.morph.parse(self.midname)[0]
        self.otch_d = self.otch_w.inflect({'sing', 'datv'}).word.capitalize()
        self.context = {
            'fam': self.surname,
            'name': self.name,
            'otch': self.midname,
            'spec': self.speciality_a,
            'fio': self.fio,
            'fam_d': self.fam_d,
            'name_d': self.name_d,
            'otch_d': self.otch_d
        }

        self.word_doc = DocxTemplate('docs\СправкаВызов.docx')
        self.word_doc.render(self.context)
        self.word_doc.save('res\СправкаВызов_рез.docx')
        self.conn.close()
        self.msg = QMessageBox()
        self.msg.setWindowTitle("Уведомление")
        self.msg.setText("Отчёт сформирован")
        self.msg.setIcon(QMessageBox.Warning)
        self.msg.exec_()




class Window(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(1126, 826)
        self.stacked_widget = QtWidgets.QStackedWidget()
        self.setCentralWidget(self.stacked_widget)

        self.m_pages = {}

        self.register(MainWindow(), "main")
        self.register(Abiturient(), "Abiturient")
        self.register(Slushatel(), "Slushatel")
        self.register(Create_anketa(), "Create_anketa")
        self.register(Prosmotret_abiturientov(), "Prosmotret_abiturientov")
        # self.register(Kafedra(), "Kafedra")
        # self.register(God_nabora(), "God_nabora")
        self.register(Anketa_abiturienta(), "Anketa_abiturienta")
        self.register(New_Year(), "New_Year")
        # self.register(Spisok_kursa(),"Spisok_kursa")
        # self.register(Anketa_slushatela(),"Anketa_slushatela")
        self.register(Otchet(),"Otchet")
        self.goto("main")

    def register(self, widget, name):
        self.m_pages[name] = widget
        self.stacked_widget.addWidget(widget)
        if isinstance(widget, PageWindow):
            widget.gotoSignal.connect(self.goto)

    @QtCore.pyqtSlot(str)
    def goto(self, name):
        if name in self.m_pages:
            widget = self.m_pages[name]
            self.stacked_widget.setCurrentWidget(widget)
            self.setWindowTitle(widget.windowTitle())


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    w = Window()
    w.show()
    sys.exit(app.exec_())
